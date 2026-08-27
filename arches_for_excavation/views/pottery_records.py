import json
from io import BytesIO
from collections import defaultdict
from functools import lru_cache

from django.http import HttpResponse, JsonResponse
from django.views import View
from arches.app.models.models import Node
from arches.app.models.tile import Tile
from django.db import transaction
from openpyxl import Workbook
from arches_for_excavation.utils.pottery.common import normalize_tile_value
from arches_for_excavation.utils.pottery.constants import POTTERY_RECORD_TYPES
from arches_for_excavation.utils.pottery.concept_lookup import (
    apply_dictionary_alias,
    get_dictionary_options,
    get_invalid_dictionary_nodes,
    resolve_dictionary_value,
    validate_and_prepare_dictionary_records,
)
from arches_for_excavation.utils.pottery.generic_record_parser import (
    create_pottery_record,
    parse_pottery_record_workbook,
)

POTTERY_REPORT_TYPES = {
    record_type: {
        "label": record_config["label"],
        "graph_id": record_config["graph_id"],
        "columns": record_config["columns"],
    }
    for record_type, record_config in POTTERY_RECORD_TYPES.items()
}

# Labels are resolved from the copied Pottery Record graphs by their visible
# node names. Keep these names aligned with the five current models.
POTTERY_REPORT_FIELDS = (
    ("formNo", "Form ID"), ("pNo", "P Number"), ("count", "Quantity"),
    ("mgNo", "MG no"), ("subcategory", "Subcategory"),
    ("vesselForm", "Vessel Form"), ("type", "Type"),
    ("typeUncertain", "Type Uncertainty"), ("vesselPart", "Vessel Part"),
    ("morphology", "Morphology"), ("surfaceTreatment", "Surface Treatment"),
    ("stateOfPreservation", "State Of Preservation"), ("chronology", "Period"),
    ("chronologyUncertain", "Uncertain"), ("provenance", "Provenance"),
    ("provenanceUncertain", "Provenance Uncertainty"), ("author", "Author"),
    ("drawn", "Drawing"), ("photo", "Photo"),
    ("specialFindId", "Special Find ID"), ("comment", "Comment"),
)

@lru_cache(maxsize=None)
def get_pottery_report_schema(graph_id):
    """Resolve report fields by name, selecting Comment from Pottery Study."""
    nodes = list(Node.objects.filter(graph_id=graph_id))
    nodes_by_name = {}
    pottery_study_group_id = next(
        (
            str(node.nodeid)
            for node in nodes
            if node.name == "Pottery Study" and node.datatype == "semantic"
        ),
        "",
    )
    for node in nodes:
        if node.name == "Comment":
            # Do not show the Comment in the nested chronology Dating card.
            # The report needs the Comment in the Pottery Study card.
            if str(node.nodegroup_id) != pottery_study_group_id:
                continue
        nodes_by_name.setdefault(node.name, str(node.nodeid))

    relation_node_id = nodes_by_name.get("Pottery Collection")
    field_map = {key: nodes_by_name.get(label) for key, label in POTTERY_REPORT_FIELDS}
    missing_fields = [key for key, node_id in field_map.items() if not node_id]
    return relation_node_id, field_map, missing_fields


def get_record_config_response(record_type):
    record_config = POTTERY_RECORD_TYPES.get(record_type)

    if not record_config:
        return None, JsonResponse(
            {
                "status": "error",
                "message": f"Unknown pottery record type: {record_type}",
            },
            status=404,
        )

    return record_config, None


def get_dictionary_options_by_field(record_config):
    return {
        field: get_dictionary_options(dictionary)
        for field, dictionary in record_config.get("dictionary_fields", {}).items()
    }


def build_preview_record(record_config, record):
    dictionary_fields = record_config.get("dictionary_fields", {})
    prepared_record, missing_values = validate_and_prepare_dictionary_records(
        record_config, [record]
    )
    prepared_record = prepared_record[0]
    missing_by_field = {value["field"]: value for value in missing_values}
    preview_record = {
        **record,
        "_dictionaryValues": {},
        "_missingDictionaryFields": [],
        "_cellErrors": {},
    }

    for field, dictionary in dictionary_fields.items():
        raw_value = record.get(field)
        value_id = prepared_record.get(field, "")
        # The workflow currently has a single-value dropdown. Keep an
        # automatically resolved concept-list in the source field; commit
        # resolves and writes all of its values without user intervention.
        preview_record["_dictionaryValues"][field] = value_id if isinstance(value_id, str) else ""
        missing_value = missing_by_field.get(field)
        if raw_value and missing_value:
            preview_record["_missingDictionaryFields"].append(field)
            preview_record["_cellErrors"][field] = (
                f"Value '{missing_value['value']}' does not exist in {dictionary}."
            )

    return preview_record


def build_missing_dictionary_values(record_config, records):
    _, missing_dictionary_values = validate_and_prepare_dictionary_records(
        record_config,
        records,
    )

    return missing_dictionary_values


def serialize_created_record(record, resource):
    resource_id = str(resource.resourceinstanceid)

    return {
        **record,
        "resourceId": resource_id,
        "resourceLink": f"/report/{resource_id}",
    }


class PotteryRecordTemplateView(View):
    """Download an XLSX template for one Pottery Record type."""

    def get(self, request, record_type):
        record_config, error_response = get_record_config_response(record_type)
        if error_response:
            return error_response

        workbook = Workbook()
        template_sheet = workbook.active
        template_sheet.title = "Template"
        dictionary_options = {
            field: get_dictionary_options(dictionary)
            for field, dictionary in record_config.get("dictionary_fields", {}).items()
        }
        template_sheet.append(["Pottery Trench", "Trench 1"])
        template_sheet.append(["Context", "Context 1"])
        template_sheet.append([])

        field_sources = [
            (field, source_keys[0])
            for field, source_keys in record_config["excel_fields"].items()
            if field in record_config["fields"]
        ]
        template_sheet.append(["form_no", "p_no", "count", *(source for _, source in field_sources)])
        sample_values = ["13038-22", "P1", 1]
        for field, _ in field_sources:
            if field in dictionary_options:
                options = dictionary_options[field]
                sample_values.append(options[0]["label"] if options else "")
            elif field in {"typeUncertain", "chronologyUncertain", "provenanceUncertain"}:
                sample_values.append("?")
            elif field == "drawn":
                sample_values.append("Yes")
            elif field == "photo":
                sample_values.append("No")
            else:
                sample_values.append("")
        template_sheet.append(sample_values)
        template_sheet.freeze_panes = "A5"

        # Keep every dictionary in its own column. This makes the sheet
        # usable as a quick reference while the user fills in the template:
        # row 1 is the dictionary name and the rows below contain its values.
        dictionary_sheet = workbook.create_sheet("Dictionary values")
        for column_index, (field, dictionary) in enumerate(
            record_config.get("dictionary_fields", {}).items(),
            start=1,
        ):
            dictionary_sheet.cell(row=1, column=column_index, value=dictionary)
            for row_index, option in enumerate(dictionary_options[field], start=2):
                dictionary_sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=option["label"],
                )
        dictionary_sheet.freeze_panes = "A2"

        for sheet in (template_sheet, dictionary_sheet):
            for column in sheet.columns:
                letter = column[0].column_letter
                sheet.column_dimensions[letter].width = min(
                    max(len(str(cell.value or "")) for cell in column) + 2,
                    50,
                )

        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f"attachment; filename=pottery-record-{record_type}-template.xlsx"
        )
        return response


class PotteryRecordPreviewView(View):
    def post(self, request, record_type):
        record_config, error_response = get_record_config_response(record_type)
        if error_response:
            return error_response

        pottery_collection_resource_id = request.POST.get("potteryCollectionResourceId")
        uploaded_file = request.FILES.get("file")

        if not pottery_collection_resource_id:
            return JsonResponse({"status": "error", "message": "Missing potteryCollectionResourceId."}, status=400)

        if not uploaded_file:
            return JsonResponse({"status": "error", "message": "Missing file."}, status=400)

        invalid_dictionary_nodes = get_invalid_dictionary_nodes(record_config)

        sheet_name, records = parse_pottery_record_workbook(uploaded_file, record_type)
        preview_records = [
            build_preview_record(record_config, record)
            for record in records
        ]

        return JsonResponse({
            "status": "preview",
            "message": f"Parsed {record_config['label']} records.",
            "potteryCollectionResourceId": pottery_collection_resource_id,
            "sheet": sheet_name,
            "parsed": len(preview_records),
            "records": preview_records,
            "recordType": record_type,
            "label": record_config["label"],
            "columns": [
                column
                for column in record_config.get("columns", [])
                if column.get("key") != "resourceLink"
            ],
            "dictionaryFields": record_config.get("dictionary_fields", {}),
            "dictionaryOptionsByField": get_dictionary_options_by_field(record_config),
            "missingDictionaryValues": build_missing_dictionary_values(record_config, records),
            "invalidDictionaryNodes": invalid_dictionary_nodes,
        })


class PotteryRecordCommitView(View):
    def post(self, request, record_type):
        record_config, error_response = get_record_config_response(record_type)
        if error_response:
            return error_response

        try:
            payload = json.loads(request.body.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            return JsonResponse({"status": "error", "message": "Invalid JSON payload."}, status=400)

        pottery_collection_resource_id = payload.get("potteryCollectionResourceId")
        records = payload.get("records") or []

        if not pottery_collection_resource_id:
            return JsonResponse({"status": "error", "message": "Missing potteryCollectionResourceId."}, status=400)

        if not records:
            return JsonResponse({"status": "error", "message": "No records to import."}, status=400)

        invalid_dictionary_nodes = get_invalid_dictionary_nodes(record_config)
        if invalid_dictionary_nodes:
            return JsonResponse(
                {
                    "status": "error",
                    "message": "Dictionary nodes must use concept or concept-list datatype.",
                    "invalidDictionaryNodes": invalid_dictionary_nodes,
                    "recordType": record_type,
                    "label": record_config["label"],
                },
                status=400,
            )

        normalized_records = []
        dictionary_fields = record_config.get("dictionary_fields", {})

        for record in records:
            normalized_record = {
                key: value
                for key, value in record.items()
                if not key.startswith("_")
            }
            dictionary_values = record.get("_dictionaryValues") or {}

            for field in dictionary_fields:
                if dictionary_values.get(field):
                    normalized_record[field] = dictionary_values[field]

            if normalized_record.get("count") == "":
                normalized_record["count"] = None
            elif isinstance(normalized_record.get("count"), str) and normalized_record["count"].isdigit():
                normalized_record["count"] = int(normalized_record["count"])

            normalized_records.append(normalized_record)

        prepared_records, missing_dictionary_values = validate_and_prepare_dictionary_records(
            record_config,
            normalized_records,
        )

        if missing_dictionary_values:
            return JsonResponse(
                {
                    "status": "error",
                    "message": "Missing dictionary values.",
                    "missingDictionaryValues": missing_dictionary_values,
                    "recordType": record_type,
                    "label": record_config["label"],
                },
                status=400,
            )

        created_records = []
        with transaction.atomic():
            for record, prepared_record in zip(normalized_records, prepared_records):
                resource = create_pottery_record(
                    record_type,
                    prepared_record,
                    pottery_collection_resource_id,
                )
                display_record = dict(record)
                for field in dictionary_fields:
                    if prepared_record.get(field):
                        display_record[field] = normalize_tile_value(prepared_record[field])

                created_records.append(serialize_created_record(display_record, resource))

        return JsonResponse({
            "status": "success",
            "message": f"Created {record_config['label']} records.",
            "potteryCollectionResourceId": pottery_collection_resource_id,
            "created": len(created_records),
            "records": created_records[:20],
            "recordType": record_type,
            "label": record_config["label"],
            "columns": record_config.get("columns", []),
        })


class PotteryRecordImportView(View):
    record_type = None

    def post(self, request):
        record_config = POTTERY_RECORD_TYPES.get(self.record_type)
        pottery_collection_resource_id = request.POST.get("potteryCollectionResourceId")
        uploaded_file = request.FILES.get("file")

        if not record_config:
            return JsonResponse(
                {
                    "status": "error",
                    "message": f"Unknown pottery record type: {self.record_type}",
                },
                status=404,
            )

        if not pottery_collection_resource_id:
            return JsonResponse({"status": "error", "message": "Missing potteryCollectionResourceId."}, status=400)

        if not uploaded_file:
            return JsonResponse({"status": "error", "message": "Missing file."}, status=400)

        sheet_name, records = parse_pottery_record_workbook(uploaded_file, self.record_type)
        prepared_records, missing_dictionary_values = validate_and_prepare_dictionary_records(
            record_config,
            records,
        )

        if missing_dictionary_values:
            return JsonResponse(
                {
                    "status": "error",
                    "message": "Missing dictionary values.",
                    "missingDictionaryValues": missing_dictionary_values,
                    "recordType": self.record_type,
                    "label": record_config["label"],
                    "sheet": sheet_name,
                },
                status=400,
            )

        invalid_dictionary_nodes = get_invalid_dictionary_nodes(record_config)
        if invalid_dictionary_nodes:
            return JsonResponse(
                {
                    "status": "error",
                    "message": "Dictionary nodes must use concept or concept-list datatype.",
                    "invalidDictionaryNodes": invalid_dictionary_nodes,
                    "recordType": self.record_type,
                    "label": record_config["label"],
                    "sheet": sheet_name,
                },
                status=400,
            )

        created_records = []

        with transaction.atomic():
            for record, prepared_record in zip(records, prepared_records):
                resource = create_pottery_record(
                    self.record_type,
                    prepared_record,
                    pottery_collection_resource_id,
                )
                resource_id = str(resource.resourceinstanceid)

                created_records.append({
                    **record,
                    "resourceId": resource_id,
                    "resourceLink": f"/report/{resource_id}",
                })

        return JsonResponse({
            "status": "success",
            "message": f"Created {record_config['label']} records.",
            "potteryCollectionResourceId": pottery_collection_resource_id,
            "sheet": sheet_name,
            "created": len(created_records),
            "records": created_records[:20],
            "recordType": self.record_type,
            "label": record_config["label"],
            "columns": record_config.get("columns", []),
        })

class AmphoraeRecordImportView(PotteryRecordImportView):
    record_type = "amphorae"

class StorageVesselRecordImportView(PotteryRecordImportView):
    record_type = "storage-vessel"

class TableWareRecordImportView(PotteryRecordImportView):
    record_type = "table-ware"

class PotteryRecordsForCollectionView(View):
    def get(self, request, collection_resource_id, record_type):
        report_config = POTTERY_REPORT_TYPES.get(record_type)

        if not report_config:
            return JsonResponse(
                {
                    "status": "error",
                    "message": f"Unknown pottery record type: {record_type}",
                },
                status=404,
            )

        relation_node_id, field_map, missing_fields = get_pottery_report_schema(
            report_config["graph_id"]
        )
        if not relation_node_id or missing_fields:
            return JsonResponse(
                {
                    "status": "error",
                    "message": "The Pottery Record graph is missing required report fields.",
                    "missingFields": missing_fields,
                    "recordType": record_type,
                },
                status=500,
            )

        # Filter the relationship in PostgreSQL instead of examining every
        # record of this type in Python. ``data__contains`` works with the
        # JSON array stored by the resource-instance datatype.
        relation_tiles = Tile.objects.filter(
            resourceinstance__graph_id=report_config["graph_id"],
            nodegroup_id=relation_node_id,
            data__contains={
                relation_node_id: [{"resourceId": str(collection_resource_id)}]
            },
        )

        # Fetch every tile for the matching records in one query. The former
        # implementation issued one additional query for every record (N+1),
        # which made collections with many fragments progressively slower.
        resource_ids = list(
            relation_tiles.values_list("resourceinstance_id", flat=True).distinct()
        )
        tiles_by_resource_id = defaultdict(list)
        for tile in Tile.objects.filter(resourceinstance_id__in=resource_ids).values(
            "resourceinstance_id", "data"
        ):
            tiles_by_resource_id[str(tile["resourceinstance_id"])].append(
                tile["data"] or {}
            )

        records = []
        for resource_id in resource_ids:

            row = {
                "resourceId": str(resource_id),
                "resourceLink": f"/report/{resource_id}",
            }

            for output_key in field_map.keys():
                row[output_key] = ""

            for data in tiles_by_resource_id[str(resource_id)]:
                for output_key, node_id in field_map.items():
                    if node_id in data:
                        display_value = normalize_tile_value(data.get(node_id))
                        if output_key == "vesselPart" and row[output_key] and display_value:
                            row[output_key] = f"{row[output_key]}, {display_value}"
                        else:
                            row[output_key] = display_value
            source_sheet = row.get("sourceSheet", "")
            source_row = row.get("sourceRow", "")

            if source_sheet and source_row:
                row["source"] = f"{source_sheet} #{source_row}"
            elif source_sheet:
                row["source"] = source_sheet
            else:
                row["source"] = ""            
            records.append(row)

        return JsonResponse({
            "status": "success",
            "collectionResourceId": str(collection_resource_id),
            "recordType": record_type,
            "label": report_config["label"],
            "columns": report_config["columns"],
            "count": len(records),
            "records": records,
        })
