import json
import re
from uuid import uuid4

from django.db import transaction
from django.http import JsonResponse
from django.views import View
from openpyxl import load_workbook

from arches.app.models.models import Node, Resource, Value
from arches.app.models.tile import Tile
from arches_for_excavation.utils.pottery.concept_lookup import resolve_dictionary_value
from arches_for_excavation.utils.pottery.constants import POTTERY_DICTIONARY_CHRONOLOGY

POTTERY_GRAPH_ID = "32a4c0b9-ab8c-47a0-a42f-99cd3ad392fe"
CONTEXT_NODE_ID = "622addb9-60c1-498c-ab40-bef9ded91f2f"
ARCHAEOLOGICAL_REMAINS_NODE_ID = "6b77bb10-1d42-445f-bbc6-dc2e5db3f129"
SPECIAL_FINDS_NODE_ID = "e912a3bd-dfa8-485a-9a8b-0ed952aafbd9"
POTTERY_FRAGMENTS_NODEGROUP_ID = "8f7a5ca4-9c49-405d-9a08-a8debb13a9ec"
RESOURCE_FORM_ID_NODE_ID = "25e31613-69ac-45ce-a6db-a15239de70a4"
LAST_SHRED_NO_NODE_ID = "51618119-e3a6-4bcb-bb93-ce2987f7ac56"
POTTERY_TYPE_NODE_ID = "3bc235a3-2240-4e94-b8af-f4c70ee13af0"
UNDIAGNOSTIC_NODE_ID = "edecb9f5-4b8e-4d6f-890b-76f8a3521b41"
NO_MATERIAL_NODE_ID = "3d57d956-38a5-4982-a0f6-d8fb388e1cb9"
CATEGORY_REMARKS_NODE_ID = "3c371503-9028-464a-8b85-53a43c853781"
CATEGORY_CHRONOLOGY_NODEGROUP_ID = "13c63c03-ffc3-455a-a1ce-23082b4111e8"
CATEGORY_PERIOD_NODE_ID = "ab05ac4b-4fd8-4eb9-9549-9d4a2a86893c"
CONTAINS_SPECIAL_FINDS_NODE_ID = "f0930ca8-a24f-458c-a27e-463601ca574a"
DIAGNOSTIC_NODE_ID = "99affc33-fc6e-4fee-9ac4-e2d4f13087cc"
FINDS_DICTIONARY_ID = "401ad5c8-4a6d-40fc-a9cd-61e1d953f13e"
FIELD_REMAINS_LABEL_MAP = {
    "b_presence": "Bones",
    "b_objects_presence": "Bricks",
    "c_presence": "Pottery",
    "g_presence": "Glass",
    "m_presence": "Metals",
    "pp_presence": "Pipes",
    "pl_presence": "Plasters",
    "sp_presence": "Other",
    "sh_presence": "Shells",
    "s_presence": "Stones",
    "tr_presence": "Terracotta",
    "t_presence": "Tiles",
    "v_presence": "Varia",
}


# The Pottery Collection graph uses the PAC category concepts configured on
# its Pottery Type node. These are the stable PAC IDs for the category
# columns in the simplified collection workbook.
POTTERY_TYPE_PAC_IDS = {
    "Amphorae": "Q969",
    "Table Ware": "Q937",
    "Kitchen Ware": "Q970",
    "Plain Ware": "Q938",
    "Storage Vessel": "Q971",
    "Lamp": "Q924",
}
PAC_ENTITY_URL = "https://pac.cenagis.edu.pl/entity/"

def normalize_header(header):
    normalized = re.sub(r"[\s\-/]+", "_", str(header or "").replace("\x00", "").strip().lower())
    normalized = re.sub(r"[^a-z0-9_]+", "", normalized)
    normalized = re.sub(r"_+", "_", normalized)
    return normalized.strip("_")


def clean_workbook_cell(value):
    if value is None:
        return ""

    if isinstance(value, bool):
        return "true" if value else "false"

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def parse_pottery_workbook(uploaded_file):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)

    for sheet in workbook.worksheets[:1]:
        rows = list(sheet.iter_rows(values_only=True))
        header_index = None

        for index, row in enumerate(rows):
            if any(clean_workbook_cell(value) for value in row):
                header_index = index
                break

        if header_index is None:
            continue

        headers = [clean_workbook_cell(value) for value in rows[header_index]]
        normalized_headers = [normalize_header(header) for header in headers]
        data_rows = []
        preview_rows = []

        for row in rows[header_index + 1:]:
            cleaned_values = [clean_workbook_cell(value) for value in row[:len(headers)]]

            if not any(cleaned_values):
                continue

            row_data = {}

            for index, header in enumerate(normalized_headers):
                if header:
                    row_data[header] = cleaned_values[index] if index < len(cleaned_values) else ""

            data_rows.append(row_data)

            if len(preview_rows) < 5:
                preview_rows.append(cleaned_values)

            # A collection workflow is deliberately limited to one source row.
            break

        if headers:
            return {
                "sheet": sheet.title,
                "headers": headers,
                "rows": data_rows,
                "previewRows": preview_rows,
                "rowCount": len(data_rows),
            }

    raise ValueError("XLSX does not contain a usable table.")


def localized_string(value):
    return {
        "en": {
            "value": value or "",
            "direction": "ltr",
        }
    }


class PotteryImportWorkbookPreviewView(View):
    def post(self, request):
        uploaded_file = request.FILES.get("file")

        if not uploaded_file:
            return JsonResponse(
                {"status": "error", "message": "Missing file."},
                status=400,
            )

        if not uploaded_file.name.lower().endswith(".xlsx"):
            return JsonResponse(
                {"status": "error", "message": "Choose an XLSX file."},
                status=400,
            )

        try:
            parsed = parse_pottery_workbook(uploaded_file)
        except Exception as error:
            return JsonResponse(
                {"status": "error", "message": f"Could not read XLSX: {error}"},
                status=400,
            )

        return JsonResponse({
            "status": "success",
            "format": "xlsx",
            "encoding": "xlsx",
            "delimiter": "",
            "columnCount": len(parsed["headers"]),
            **parsed,
        })


class PotteryImportPreviewView(View):
    @staticmethod
    def _resolve_pottery_type(pottery_type):
        """Resolve the fixed category labels used by collection workbooks."""
        pottery_type_value = resolve_dictionary_value("Pottery Type", pottery_type)
        if pottery_type_value:
            return pottery_type_value

        pac_id = POTTERY_TYPE_PAC_IDS.get(pottery_type)
        if not pac_id:
            return ""

        values = Value.objects.filter(
            concept__legacyoid=f"{PAC_ENTITY_URL}{pac_id}",
            valuetype_id="prefLabel",
        ).order_by("language_id")
        english_value = values.filter(language_id="en").first()
        value = english_value or values.first()
        return str(value.valueid) if value else ""

    def _find_existing_pottery_collection(self, context_resource_id):
        tiles = Tile.objects.filter(
            nodegroup_id=CONTEXT_NODE_ID,
        )

        for tile in tiles:
            value = tile.data.get(CONTEXT_NODE_ID)

            if not isinstance(value, list):
                continue

            for relation in value:
                if relation.get("resourceId") == context_resource_id:
                    return tile.resourceinstance

        return None

    @staticmethod
    def _create_tile(resource, nodegroup_id, data, parent_tile=None):
        tile = Tile.get_blank_tile_from_nodegroup_id(
            nodegroup_id,
            resourceid=str(resource.resourceinstanceid),
        )
        tile.parenttile = parent_tile
        tile.data.update(data)
        tile.save()
        return tile

    @staticmethod
    def _merge_collection_context_values(pottery_resource, archaeological_remains, special_finds):
        """Add resolved finds to a collection that has no pottery fragments."""
        context_tile = Tile.objects.filter(
            resourceinstance_id=pottery_resource.resourceinstanceid,
            nodegroup_id=CONTEXT_NODE_ID,
        ).first()
        if context_tile is None:
            return

        context_tile.data = dict(context_tile.data or {})
        for node_id, values in (
            (ARCHAEOLOGICAL_REMAINS_NODE_ID, archaeological_remains),
            (SPECIAL_FINDS_NODE_ID, special_finds),
        ):
            if values:
                existing_values = list(context_tile.data.get(node_id) or [])
                context_tile.data[node_id] = list(dict.fromkeys([*existing_values, *values]))
        context_tile.save()

    @classmethod
    def _create_root_tile_for_node(cls, resource, node_id, value):
        """Creates a root tile using the nodegroup defined by the active graph."""
        node = Node.objects.get(nodeid=node_id)
        return cls._create_tile(
            resource,
            str(node.nodegroup_id),
            {node_id: value},
        )

    @staticmethod
    def _concept_values_from_flags(flags, suffix="_presence"):
        value_ids = []

        for source_key, concept_label in FIELD_REMAINS_LABEL_MAP.items():
            flag_key = source_key if suffix == "_presence" else source_key.replace("_presence", suffix)
            concept_value_id = resolve_dictionary_value(
                FINDS_DICTIONARY_ID,
                concept_label,
            )
            if flags.get(flag_key) and concept_value_id and concept_value_id not in value_ids:
                value_ids.append(concept_value_id)

        return value_ids

    @staticmethod
    def _number_or_none(value):
        if value in (None, ""):
            return None

        if isinstance(value, (int, float)):
            return value

        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _create_fragment(self, pottery_resource, item, missing_concepts):
        pottery_type = item.get("type", "")
        pottery_type_value = self._resolve_pottery_type(pottery_type)

        if not pottery_type_value:
            missing_concepts.append({
                "field": "potteryType",
                "value": pottery_type,
                "potteryType": pottery_type,
            })
            return False

        fragment_data = {
            POTTERY_TYPE_NODE_ID: pottery_type_value,
            NO_MATERIAL_NODE_ID: bool(item.get("noMaterial")),
            CONTAINS_SPECIAL_FINDS_NODE_ID: bool(item.get("specialFind")),
        }
        undiagnostic = self._number_or_none(item.get("undiagnostic"))
        diagnostic = self._number_or_none(item.get("diagnostic"))
        remarks = item.get("remarks")
        chronology = item.get("chronology")

        if undiagnostic is not None:
            fragment_data[UNDIAGNOSTIC_NODE_ID] = undiagnostic
        if diagnostic is not None:
            fragment_data[DIAGNOSTIC_NODE_ID] = diagnostic
        if remarks:
            fragment_data[CATEGORY_REMARKS_NODE_ID] = localized_string(remarks)
        if chronology:
            chronology_value = resolve_dictionary_value(
                POTTERY_DICTIONARY_CHRONOLOGY,
                chronology,
            )
            if chronology_value:
                category_chronology_value = chronology_value
            else:
                missing_concepts.append({
                    "field": "categoryChronology",
                    "value": chronology,
                    "potteryType": pottery_type,
                })
                category_chronology_value = ""
        else:
            category_chronology_value = ""

        fragment_tile = self._create_tile(
            pottery_resource,
            POTTERY_FRAGMENTS_NODEGROUP_ID,
            fragment_data,
        )

        if category_chronology_value:
            self._create_tile(
                pottery_resource,
                CATEGORY_CHRONOLOGY_NODEGROUP_ID,
                {CATEGORY_PERIOD_NODE_ID: [category_chronology_value]},
                parent_tile=fragment_tile,
            )

        return True

    def post(self, request):
        try:
            payload = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            return JsonResponse(
                {"status": "error", "message": "Invalid JSON payload."},
                status=400,
            )

        context_resource_id = payload.get("contextResourceId")
        context_number = payload.get("contextNumber")
        pottery_rows = payload.get("rows", [])

        if not context_resource_id:
            return JsonResponse(
                {"status": "error", "message": "Missing contextResourceId."},
                status=400,
            )

        if not isinstance(pottery_rows, list):
            return JsonResponse(
                {"status": "error", "message": "rows must be a list."},
                status=400,
            )

        if not pottery_rows:
            return JsonResponse(
                {"status": "error", "message": "No pottery rows with importable pottery data were found."},
                status=400,
            )

        source_context_numbers = {
            str(row.get("contextNo", "")).strip()
            for row in pottery_rows
            if str(row.get("contextNo", "")).strip()
        }
        if source_context_numbers != {str(context_number).strip()}:
            return JsonResponse(
                {
                    "status": "error",
                    "message": "Every imported row must belong to the selected Context.",
                    "selectedContextNumber": context_number,
                    "sourceContextNumbers": sorted(source_context_numbers),
                },
                status=400,
            )

        form_ids = {
            str(row.get("formId", "")).strip()
            for row in pottery_rows
            if str(row.get("formId", "")).strip()
        }
        last_shred_nos = {
            str(row.get("lastShredNo", "")).strip()
            for row in pottery_rows
            if str(row.get("lastShredNo", "")).strip()
        }
        if len(form_ids) > 1 or len(last_shred_nos) > 1:
            return JsonResponse(
                {
                    "status": "error",
                    "message": "One Pottery Collection import must contain one Form ID and one Last Shred No.",
                    "formIds": sorted(form_ids),
                    "lastShredNos": sorted(last_shred_nos),
                },
                status=400,
            )

        last_shred_no_value = None
        if last_shred_nos:
            last_shred_no_value = self._number_or_none(next(iter(last_shred_nos)))
            if last_shred_no_value is None:
                return JsonResponse(
                    {
                        "status": "error",
                        "message": "Last Shred No must be a number.",
                        "lastShredNo": next(iter(last_shred_nos)),
                    },
                    status=400,
                )

        related_context_value = [{
            "resourceId": context_resource_id,
            "ontologyProperty": "",
            "inverseOntologyProperty": "",
            "resourceXresourceId": str(uuid4()),
        }]
        archaeological_remains = []
        special_finds = []
        for row in pottery_rows:
            for value_id in self._concept_values_from_flags(row.get("fieldRemains", {})):
                if value_id not in archaeological_remains:
                    archaeological_remains.append(value_id)
            for value_id in self._concept_values_from_flags(
                row.get("specialFinds", {}),
                suffix="_special_finds",
            ):
                if value_id not in special_finds:
                    special_finds.append(value_id)

        with transaction.atomic():
            pottery_resource = self._find_existing_pottery_collection(context_resource_id)
            context_resource = Resource.objects.get(resourceinstanceid=context_resource_id)
            if str(context_resource.graph_id) != "2c536779-d3e6-43ef-bc0c-cd4d97dc8c6c":
                return JsonResponse(
                    {
                        "status": "error",
                        "message": "The new Pottery Collection can only be linked to an (O) Context resource.",
                    },
                    status=400,
                )

            is_new_collection = pottery_resource is None
            should_create_fragments = is_new_collection
            if not is_new_collection:
                # Re-importing never duplicates fragment cards, but it does
                # merge the source-level Archaeological Remains and Special Finds.
                should_create_fragments = not Tile.objects.filter(
                    resourceinstance_id=pottery_resource.resourceinstanceid,
                    nodegroup_id=POTTERY_FRAGMENTS_NODEGROUP_ID,
                ).exists()
                self._merge_collection_context_values(
                    pottery_resource, archaeological_remains, special_finds
                )
            else:
                pottery_resource = Resource.objects.create(graph_id=POTTERY_GRAPH_ID)
                context_data = {CONTEXT_NODE_ID: related_context_value}
                if archaeological_remains:
                    context_data[ARCHAEOLOGICAL_REMAINS_NODE_ID] = archaeological_remains
                if special_finds:
                    context_data[SPECIAL_FINDS_NODE_ID] = special_finds
                self._create_tile(pottery_resource, CONTEXT_NODE_ID, context_data)

            if is_new_collection and form_ids:
                self._create_root_tile_for_node(
                    pottery_resource,
                    RESOURCE_FORM_ID_NODE_ID,
                    next(iter(form_ids)),
                )
            if is_new_collection and last_shred_nos:
                self._create_root_tile_for_node(
                    pottery_resource,
                    LAST_SHRED_NO_NODE_ID,
                    last_shred_no_value,
                )

            missing_concepts = []
            created_fragments = 0
            for row in (pottery_rows if should_create_fragments else []):
                for item in row.get("pottery", []):
                    if self._create_fragment(pottery_resource, item, missing_concepts):
                        created_fragments += 1

        return JsonResponse({
            "status": "success",
            "message": "Created Pottery Collection resource." if is_new_collection else "Updated Pottery Collection source fields.",
            "potteryCollectionResourceId": str(pottery_resource.resourceinstanceid),
            "contextResourceId": context_resource_id,
            "contextNumber": context_number,
            "formId": next(iter(form_ids), ""),
            "lastShredNo": next(iter(last_shred_nos), ""),
            "rowCount": len(pottery_rows),
            "fragmentCount": created_fragments,
            "rows": pottery_rows,
            "archaeologicalRemains": archaeological_remains,
            "specialFinds": special_finds,
            "missingConcepts": missing_concepts,
            "created": is_new_collection,
            "updated": not is_new_collection,
        })
