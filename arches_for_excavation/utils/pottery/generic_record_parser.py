import re
from uuid import uuid4

from openpyxl import load_workbook
from arches.app.models.models import Node, Resource
from arches.app.models.tile import Tile
from arches.app.utils.date_utils import ExtendedDateFormat

from arches_for_excavation.utils.pottery.common import (
    clean_cell,
    create_tile_for_node,
    decode_vessel_part,
    get_value,
    localized_string,
    normalize_header,
    parse_p_no_and_count,
    to_boolean,
)
from arches_for_excavation.utils.pottery.concept_lookup import format_concept_tile_value
from arches_for_excavation.utils.pottery.constants import POTTERY_RECORD_TYPES


DATA_SLOT_COUNT = 36
BOOLEAN_RECORD_FIELDS = {
    "drawn",
    "photo",
    "typeUncertain",
    "chronologyUncertain",
    "provenanceUncertain",
}

SPECIAL_CHRONOLOGIES = {
    "1st 2nd c ce": ("1st c. CE|2nd c. CE", None),
}


PARTIAL_CENTURY_PATTERN = re.compile(
    r"^(?P<part>early|mid(?:dle)?|late|end|(?:first|1st) half|(?:second|2nd) half)"
    r"(?: of)?(?: the)? (?P<century>\d+)(?:st|nd|rd|th)?"
    r"\s*(?:c(?:entury)?)?\s*(?P<era>ad|ce|bc|bce)$",
    flags=re.IGNORECASE,
)


def _ordinal_suffix(number):
    if 10 < number % 100 < 14:
        return "th"
    return {1: "st", 2: "nd", 3: "rd"}.get(number % 10, "th")


def normalize_partial_century(value):
    """Return the broad period and exact EDTF years for a part of a century."""
    raw = clean_cell(value)
    normalized = re.sub(r"[.]", "", raw)
    normalized = re.sub(r"\s*[-–—]\s*", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    match = PARTIAL_CENTURY_PATTERN.fullmatch(normalized)
    if not match:
        return None

    part = match.group("part").casefold()
    century = int(match.group("century"))
    era = match.group("era").upper()
    offsets = {
        "early": (1, 25), "mid": (25, 50), "middle": (25, 50),
        "late": (75, 100), "end": (75, 100),
        "first half": (0, 50), "1st half": (0, 50),
        "second half": (50, 100), "2nd half": (50, 100),
    }
    start_offset, end_offset = offsets[part]
    period_era = "CE" if era in {"AD", "CE"} else "BCE"
    period = f"{century}{_ordinal_suffix(century)} c. {period_era}"
    if period_era == "CE":
        base = (century - 1) * 100
        return period, (base + start_offset, base + end_offset)

    # 1 BCE is astronomical year 0 in EDTF.
    century_start = 1 - (century * 100)
    if part == "early":
        return period, (century_start, century_start + 24)
    if part in {"mid", "middle"}:
        return period, (century_start + 25, century_start + 49)
    if part in {"late", "end"}:
        return period, (century_start + 75, century_start + 99)
    if part in {"first half", "1st half"}:
        return period, (century_start, century_start + 49)
    return period, (century_start + 50, century_start + 99)


def normalize_chronology(value):
    raw = clean_cell(value)
    normalized = re.sub(r"[.]", "", raw)
    normalized = re.sub(r"\s*[-–—]\s*", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).casefold()
    direct_match = SPECIAL_CHRONOLOGIES.get(normalized)
    if direct_match:
        return (*direct_match, raw if direct_match[1] else "")
    partial_century = normalize_partial_century(raw)
    if partial_century:
        periods, boundaries = partial_century
        return periods, boundaries, raw
    for label, (periods, boundaries) in SPECIAL_CHRONOLOGIES.items():
        if boundaries and label in normalized:
            broad_periods = re.sub(r"\([^)]*\)", "", raw)
            broad_periods = re.sub(r"\s*[-–—]\s*", "|", broad_periods).strip("| ")
            chronology = "|".join(value for value in (broad_periods, periods) if value)
            return chronology, boundaries, raw
    return raw, None, ""


def apply_chronology_normalization(record):
    periods, boundaries, source_value = normalize_chronology(record.get("chronology"))
    record["chronology"] = periods
    record["chronologyBoundaries"] = boundaries
    if source_value:
        comment = clean_cell(record.get("comment"))
        record["comment"] = "\n".join(
            value for value in (comment, f"Period: {source_value}") if value
        )


def save_chronology_boundaries(resource, chronology_node_id, boundaries):
    if not boundaries:
        return
    period_node = Node.objects.get(nodeid=chronology_node_id)
    nodes_by_name = {
        node.name.casefold(): str(node.nodeid)
        for node in Node.objects.filter(nodegroup_id=period_node.nodegroup_id)
    }
    required_names = (
        "earliest date", "latest start date", "earliest end date", "latest date",
    )
    if any(name not in nodes_by_name for name in required_names):
        return
    start_year, end_year = boundaries
    for name, year in zip(required_names, (start_year, start_year, end_year, end_year)):
        create_tile_for_node(
            resource,
            nodes_by_name[name],
            str(ExtendedDateFormat(str(year)).edtf),
        )

def format_text_tile_value(node_id, value):
    if Node.objects.get(nodeid=node_id).datatype == "string":
        return localized_string(value)
    return value
def get_record_config(record_type):
    record_config = POTTERY_RECORD_TYPES.get(record_type)

    if not record_config:
        raise ValueError(f"Unknown pottery record type: {record_type}")

    return record_config


def has_record_data(record, record_config):
    for key in record_config.get("record_data_keys", []):
        value = record.get(key)
        if value not in (None, ""):
            return True

    return False


def find_header_row(sheet, predicate):
    for row_index, row in enumerate(sheet.iter_rows(max_row=20), start=1):
        headers = [normalize_header(cell.value) for cell in row]
        if predicate(headers):
            return row_index, headers

    return None, []


def get_first_value(values, source_keys):
    return clean_cell(get_value(values, *source_keys))


def get_suffixed_value(values, source_keys, suffix, number):
    suffixed_keys = []

    for source_key in source_keys:
        suffixed_keys.append(f"{source_key}_{suffix}")
        suffixed_keys.append(f"{source_key}_0{number}")

    return clean_cell(get_value(values, *suffixed_keys))


def normalize_field_value(key, value):
    if key == "vesselPart":
        return decode_vessel_part(value)
    if key == "morphology" and clean_cell(value).casefold() == "everted flat":
        return "everted|flat"

    return clean_cell(value)


def table_p_no(values):
    return get_value(values, "p_no", "p_number")


def build_base_record(record_config, sheet_title, row_index, context="", trench="", form_no=""):
    return {
        "potteryType": record_config["label"],
        "context": context,
        "trench": trench,
        "formNo": form_no,
        "pNo": "",
        "count": None,
        "sourceSheet": sheet_title,
        "sourceRow": row_index,
    }


def get_table_metadata(sheet, header_row_index):
    """Read common Context and Pottery Trench values written above the table."""
    metadata = {"context": "", "trench": ""}
    labels = {
        "context": "context",
        "trench": "trench",
        "pottery_trench": "trench",
    }

    for row in sheet.iter_rows(max_row=header_row_index - 1):
        cells = list(row)
        for index, cell in enumerate(cells[:-1]):
            field = labels.get(normalize_header(cell.value))
            if field:
                metadata[field] = clean_cell(cells[index + 1].value)
                break

    return metadata


def parse_table_sheet(sheet, record_type):
    record_config = get_record_config(record_type)
    header_row_index, headers = find_header_row(
        sheet,
        lambda header_values: any(key in header_values for key in ("p_no", "p_number")),
    )

    if not header_row_index:
        return []

    records = []
    metadata = get_table_metadata(sheet, header_row_index)

    for row_index, row in enumerate(sheet.iter_rows(min_row=header_row_index + 1), start=header_row_index + 1):
        values = {}

        for index, cell in enumerate(row):
            if index < len(headers):
                values[headers[index]] = clean_cell(cell.value)

        raw_count = get_value(values, "quantity", "count")
        p_no, count = parse_p_no_and_count(table_p_no(values), raw_count)
        if raw_count.isdigit():
            count = int(raw_count)

        record = build_base_record(
            record_config,
            sheet.title,
            row_index,
            context=values.get("context", metadata["context"]),
            trench=values.get("trench", metadata["trench"]),
            form_no=get_value(values, "form_no", "form_id"),
        )
        record["pNo"] = p_no
        record["count"] = count

        for key, source_keys in record_config.get("excel_fields", {}).items():
            record[key] = normalize_field_value(key, get_first_value(values, source_keys))
        apply_chronology_normalization(record)

        if has_record_data(record, record_config):
            records.append(record)

    return records


def parse_data_sheet(sheet, record_type):
    record_config = get_record_config(record_type)
    header_row_index, headers = find_header_row(
        sheet,
        lambda header_values: "context_no" in header_values and any(header.startswith("no_p_") for header in header_values),
    )

    if not header_row_index:
        return []

    records = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=header_row_index + 1), start=header_row_index + 1):
        values = {}

        for index, cell in enumerate(row):
            if index < len(headers):
                values[headers[index]] = cell.value

        context = clean_cell(values.get("context_no"))
        trench = clean_cell(values.get("trench"))
        # ``id`` is often only an Excel row identifier; never use it as a Form ID.
        form_no = clean_cell(values.get("form_no"))

        for number in range(1, DATA_SLOT_COUNT + 1):
            suffix = f"{number:02d}"
            raw_p_no = get_suffixed_value(values, ["no_p"], suffix, number)
            raw_quantity = get_suffixed_value(values, ["no_quantity"], suffix, number)
            p_no, count = parse_p_no_and_count(raw_p_no, raw_quantity)

            record = build_base_record(
                record_config,
                sheet.title,
                row_index,
                context=context,
                trench=trench,
                form_no=form_no,
            )
            record["pNo"] = p_no
            record["count"] = count

            for key, source_keys in record_config.get("excel_fields", {}).items():
                value = get_suffixed_value(values, source_keys, suffix, number)
                record[key] = normalize_field_value(key, value)
            apply_chronology_normalization(record)

            if has_record_data(record, record_config):
                records.append(record)

    return records
def parse_pottery_record_workbook(uploaded_file, record_type):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    record_config = get_record_config(record_type)
    preferred_sheets = record_config.get("sheet_priority", ["Template", "Table", "Data"])
    sheet_names = list(dict.fromkeys(preferred_sheets + workbook.sheetnames))

    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        records = parse_data_sheet(sheet, record_type) if sheet_name == "Data" else []
        if not records:
            records = parse_table_sheet(sheet, record_type)
        if records:
            return sheet_name, records

    return "", []

    return "", []


def create_pottery_record(record_type, record, pottery_collection_resource_id):
    record_config = get_record_config(record_type)
    dictionary_fields = record_config.get("dictionary_fields", {})
    resource = Resource.objects.create(graph_id=record_config["graph_id"])

    related_collection_value = [{
        "resourceId": pottery_collection_resource_id,
        "ontologyProperty": "",
        "inverseOntologyProperty": "",
        "resourceXresourceId": str(uuid4()),
    }]

    create_tile_for_node(
        resource,
        record_config["related_collection_node_id"],
        related_collection_value,
    )
    pottery_type_node_id = record_config.get("pottery_type_node_id")
    pottery_type_value = record.get("potteryType")
    if pottery_type_node_id:
        if "potteryType" in dictionary_fields and pottery_type_value:
            create_tile_for_node(
                resource,
                pottery_type_node_id,
                format_concept_tile_value(pottery_type_node_id, pottery_type_value),
            )
        else:
            create_tile_for_node(
                resource,
                pottery_type_node_id,
                localized_string(pottery_type_value),
            )

    nodes_by_id = {
        str(node.nodeid): node
        for node in Node.objects.filter(nodeid__in=record_config.get("fields", {}).values())
    }
    tile_data_by_nodegroup = {}

    for key, node_id in record_config.get("fields", {}).items():
        value = record.get(key)
        tile_value = None

        if key == "count":
            if value is not None:
                tile_value = value
        elif key in BOOLEAN_RECORD_FIELDS:
            if value not in (None, ""):
                tile_value = to_boolean(value)
        elif key in dictionary_fields:
            if value:
                tile_value = format_concept_tile_value(node_id, value)
        elif value not in (None, ""):
            tile_value = format_text_tile_value(node_id, value)

        if tile_value is None:
            continue

        node = nodes_by_id[str(node_id)]
        nodegroup_id = str(node.nodegroup_id)
        tile_data_by_nodegroup.setdefault(nodegroup_id, {})[str(node_id)] = tile_value

    for nodegroup_id, data in tile_data_by_nodegroup.items():
        tile = Tile.get_blank_tile_from_nodegroup_id(
            nodegroup_id,
            resourceid=str(resource.resourceinstanceid),
        )
        tile.data = data
        tile.save()

    save_chronology_boundaries(
        resource,
        record_config["fields"]["chronology"],
        record.get("chronologyBoundaries"),
    )

    if record_config.get("source_sheet_node_id"):
        create_tile_for_node(
            resource,
            record_config["source_sheet_node_id"],
            localized_string(record.get("sourceSheet")),
        )
    if record_config.get("source_row_node_id"):
        create_tile_for_node(
            resource,
            record_config["source_row_node_id"],
            record.get("sourceRow"),
        )

    return resource
