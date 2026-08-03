"""Update amphorae category remarks in existing Pottery Collections."""

from collections import Counter
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from arches.app.models.models import TileModel, Value
from arches_for_excavation.utils.pottery.common import clean_cell, localized_string


POTTERY_GRAPH_ID = "32a4c0b9-ab8c-47a0-a42f-99cd3ad392fe"
FORM_ID_NODE_ID = "25e31613-69ac-45ce-a6db-a15239de70a4"
FRAGMENT_NODEGROUP_ID = "8f7a5ca4-9c49-405d-9a08-a8debb13a9ec"
POTTERY_TYPE_NODE_ID = "3bc235a3-2240-4e94-b8af-f4c70ee13af0"
CATEGORY_REMARKS_NODE_ID = "3c371503-9028-464a-8b85-53a43c853781"
PAC_AMPHORAE_URL = "https://pac.cenagis.edu.pl/entity/Q969"

FORM_ID_HEADERS = ("Form ID", "Form_ID")
REMARKS_HEADER = "Remarks (from Pottery Category Form)"


class Command(BaseCommand):
    help = "Update amphorae Pottery Category remarks from an Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True)
        parser.add_argument("--sheet", help="Sheet name; defaults to the first sheet.")
        parser.add_argument(
            "--apply",
            action="store_true",
            help="Save updates. Without this flag the command is a dry-run.",
        )

    def handle(self, *args, **options):
        path = Path(options["file"])
        if not path.is_file():
            raise CommandError(f"Workbook not found: {path}")

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[options["sheet"]] if options.get("sheet") else workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [clean_cell(value) for value in next(rows)]
        except StopIteration as error:
            raise CommandError("Workbook is empty.") from error

        form_header = next((header for header in FORM_ID_HEADERS if header in headers), None)
        if form_header is None or REMARKS_HEADER not in headers:
            raise CommandError(
                "Required columns: Form ID (or Form_ID) and "
                f"{REMARKS_HEADER}."
            )

        amphorae_value_id = self._amphorae_value_id()
        apply = options["apply"]
        totals = Counter()
        self.stdout.write(f"Amphorae remarks import [{'APPLY' if apply else 'DRY-RUN'}]")

        for row_number, values in enumerate(rows, start=2):
            row = dict(zip(headers, values))
            form_id = clean_cell(row.get(form_header))
            remarks = clean_cell(row.get(REMARKS_HEADER))
            if not form_id and not remarks:
                continue
            if not form_id:
                totals["errors"] += 1
                self.stderr.write(f"  row {row_number}: missing Form ID")
                continue
            if not remarks:
                totals["skipped_empty"] += 1
                self.stdout.write(f"  row {row_number}: skipped {form_id} (empty remarks)")
                continue

            resource_ids = list(
                TileModel.objects.filter(
                    resourceinstance__graph_id=POTTERY_GRAPH_ID,
                    data__contains={FORM_ID_NODE_ID: form_id},
                )
                .values_list("resourceinstance_id", flat=True)
                .distinct()
            )
            if not resource_ids:
                totals["missing_collections"] += 1
                self.stdout.write(f"  row {row_number}: skipped {form_id} (collection not found)")
                continue
            if len(resource_ids) != 1:
                totals["errors"] += 1
                self.stderr.write(
                    f"  row {row_number}: {form_id} has {len(resource_ids)} collections"
                )
                continue

            fragments = list(
                TileModel.objects.filter(
                    resourceinstance_id=resource_ids[0],
                    nodegroup_id=FRAGMENT_NODEGROUP_ID,
                    data__contains={POTTERY_TYPE_NODE_ID: amphorae_value_id},
                )
            )
            if not fragments:
                totals["missing_fragments"] += 1
                self.stdout.write(f"  row {row_number}: skipped {form_id} (amphorae fragment not found)")
                continue
            if len(fragments) != 1:
                totals["errors"] += 1
                self.stderr.write(
                    f"  row {row_number}: {form_id} has {len(fragments)} amphorae fragments"
                )
                continue

            fragment = fragments[0]
            new_value = localized_string(remarks)
            if fragment.data.get(CATEGORY_REMARKS_NODE_ID) == new_value:
                totals["unchanged"] += 1
                self.stdout.write(f"  row {row_number}: unchanged {form_id}")
                continue

            if apply:
                with transaction.atomic():
                    fragment.data[CATEGORY_REMARKS_NODE_ID] = new_value
                    fragment.save()
                self.stdout.write(f"  row {row_number}: updated {form_id}")
            else:
                self.stdout.write(f"  row {row_number}: would update {form_id}")
            totals["updated"] += 1

        self.stdout.write("Summary:")
        for key in (
            "updated",
            "unchanged",
            "skipped_empty",
            "missing_collections",
            "missing_fragments",
            "errors",
        ):
            self.stdout.write(f"  {key}: {totals[key]}")
        if not apply:
            self.stdout.write("Dry-run only. Add --apply to save changes.")

    @staticmethod
    def _amphorae_value_id():
        values = Value.objects.filter(
            concept__legacyoid=PAC_AMPHORAE_URL,
            valuetype_id="prefLabel",
        ).order_by("language_id")
        value = values.filter(language_id="en").first() or values.first()
        if value is None:
            raise CommandError("Local PAC amphorae value Q969 was not found.")
        return str(value.valueid)
