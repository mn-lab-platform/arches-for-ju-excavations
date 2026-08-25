"""Import the simplified Pottery Collection workbook without the web workflow."""

import json
import re
from collections import Counter
from pathlib import Path
from uuid import UUID, uuid4

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from arches.app.models.models import Node, Resource, TileModel, Value
from arches.app.models.tile import Tile
from arches_for_excavation.utils.pottery.common import clean_cell, localized_string, to_boolean
from arches_for_excavation.utils.pottery.concept_lookup import (
    get_dictionary_index,
    normalize_dictionary_label,
    resolve_dictionary_value,
)
from arches_for_excavation.utils.pottery.constants import POTTERY_DICTIONARY_CHRONOLOGY


POTTERY_GRAPH_ID = "32a4c0b9-ab8c-47a0-a42f-99cd3ad392fe"
CONTEXT_GRAPH_ID = "2c536779-d3e6-43ef-bc0c-cd4d97dc8c6c"
CONTEXT_NUMBER_NODE_ID = "cf7f2532-74f3-487f-9261-bf27825fe04c"
TRENCH_GRAPH_ID = "cc91f1ff-6ea8-422c-be14-b818660f66f8"
CONTEXT_TRENCH_NODE_ID = "13e52ba6-b14d-41de-9a09-8bd1186edc10"
TRENCH_NAME_NODE_ID = "a9aff43e-0e89-4ea6-bd3d-1b56a6d756db"

# Both MAL and PAP source context identifiers end with the local Context
# Number, for example ``MAL-TT-X-1029`` and ``PAP-T-II-2111``. PAP is a
# valid excavation-area prefix, not a marker for missing data.
IGNORED_CONTEXT_PREFIXES = ("UNKNOWN-",)

# Root cards
CONTEXT_NODEGROUP_ID = "622addb9-60c1-498c-ab40-bef9ded91f2f"
CONTEXT_NODE_ID = "622addb9-60c1-498c-ab40-bef9ded91f2f"
ARCHAEOLOGICAL_REMAINS_NODE_ID = "6b77bb10-1d42-445f-bbc6-dc2e5db3f129"
SPECIAL_FINDS_NODE_ID = "e912a3bd-dfa8-485a-9a8b-0ed952aafbd9"
CONTEXT_REMARKS_NODE_ID = "150f3e4b-b190-4f35-9379-ee9a8281d4dc"
FORM_ID_NODEGROUP_ID = "25e31613-69ac-45ce-a6db-a15239de70a4"
FORM_ID_NODE_ID = "25e31613-69ac-45ce-a6db-a15239de70a4"
LAST_SHRED_NODEGROUP_ID = "51618119-e3a6-4bcb-bb93-ce2987f7ac56"
LAST_SHRED_NODE_ID = "51618119-e3a6-4bcb-bb93-ce2987f7ac56"

# General Context Chronology card
CONTEXT_CHRONOLOGY_NODEGROUP_ID = "c6e7b936-3a60-4b5e-a557-4b5f02c5a4cf"
CONTEXT_PERIOD_NODE_ID = "bdad4eee-bae5-4aaa-b2ae-4bf646f0abd5"
CONTEXT_CHRONOLOGY_COMMENT_NODE_ID = "84412f01-f54b-4cf1-8008-f627d3dccd3c"
CONTEXT_UNCERTAIN_NODE_ID = "bf2e13d5-0081-458f-bb64-b4c61ba4face"

# Repeating Pottery Fragments card and nested chronology card
FRAGMENT_NODEGROUP_ID = "8f7a5ca4-9c49-405d-9a08-a8debb13a9ec"
POTTERY_TYPE_NODE_ID = "3bc235a3-2240-4e94-b8af-f4c70ee13af0"
DIAGNOSTIC_NODE_ID = "99affc33-fc6e-4fee-9ac4-e2d4f13087cc"
UNDIAGNOSTIC_NODE_ID = "edecb9f5-4b8e-4d6f-890b-76f8a3521b41"
NO_MATERIAL_NODE_ID = "3d57d956-38a5-4982-a0f6-d8fb388e1cb9"
CONTAINS_SPECIAL_FINDS_NODE_ID = "f0930ca8-a24f-458c-a27e-463601ca574a"
CATEGORY_REMARKS_NODE_ID = "3c371503-9028-464a-8b85-53a43c853781"
FINDS_REMARKS_NODE_ID = "0e3b3752-772b-427c-bd0e-95ccfc223eca"
CATEGORY_CHRONOLOGY_NODEGROUP_ID = "13c63c03-ffc3-455a-a1ce-23082b4111e8"
CATEGORY_PERIOD_NODE_ID = "ab05ac4b-4fd8-4eb9-9549-9d4a2a86893c"
CATEGORY_CHRONOLOGY_COMMENT_NODE_ID = "c2caa4cb-ab3e-4619-808a-b2a0ff4084d6"
CATEGORY_UNCERTAIN_NODE_ID = "ca337aae-aa58-4f47-8811-c1f994252fdf"

# The collection configured on Archaeological Remains and Special Finds.
FINDS_DICTIONARY_ID = "401ad5c8-4a6d-40fc-a9cd-61e1d953f13e"
PAC_ENTITY_URL = "https://pac.cenagis.edu.pl/entity/"

# Other_* columns are ignored until their data model is agreed.
CATEGORY_TYPES = {
    "TW": "Q937",
    "A": "Q969",
    "KW": "Q970",
    "PW": "Q938",
    "SV": "Q971",
    "L": "Q924",
}

CHRONOLOGY_ABBREVIATIONS = {
    "LCL": "Late Classical",
    "LH": "Late Hellenistic",
    "EH": "Early Hellenistic",
    "MH": "Middle Hellenistic",
    "ER": "Early Roman",
}


def cell_number(value):
    raw = clean_cell(value)
    if not raw:
        return None
    try:
        number = float(raw)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def normalize_source_context(value):
    """Normalize separators, including the compact ``PAP-TTX`` form.

    Some source workbooks omit the separator between the ``T``/``TT`` trench
    marker and its Roman numeral, e.g. ``PAP-TTX-951``. Expand that form so
    it follows the canonical source syntax used by :meth:`_source_trench_name`.
    """
    context = re.sub(r"\s*[_-]\s*", "-", clean_cell(value).upper())
    return re.sub(
        r"^(PAP|MAL)-(TT|T)([IVXLCDM]+)-((?:C)?\d+)$",
        r"\1-\2-\3-\4",
        context,
    )


class SkipRow(Exception):
    """A source row intentionally excluded from this importer."""


class Command(BaseCommand):
    help = "Import simplified Pottery Collection Excel rows without the web workflow."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path to the .xlsx file inside the Arches container.")
        parser.add_argument("--context-resource-id", help="Optional fixed (O) Context resource UUID for every workbook row.")
        parser.add_argument("--context-map", help="Optional JSON map from Excel Context values to (O) Context resource UUIDs.")
        parser.add_argument(
            "--create-missing-contexts",
            action="store_true",
            help="Create a minimal (O) Context linked to its full PAP/MAL Trench when it is missing.",
        )
        parser.add_argument("--apply", action="store_true", help="Create resources. Without this flag the command only validates the workbook.")
        parser.add_argument("--sheet", help="Optional workbook sheet name. Defaults to the first sheet.")

    def handle(self, *args, **options):
        self.apply = options["apply"]
        self.context_resource_id = options.get("context_resource_id")
        self.context_map = self._load_context_map(options.get("context_map"))
        self.create_missing_contexts = options["create_missing_contexts"]
        self.created_contexts = {}
        self.trench_resource_ids = None
        if self.context_resource_id and self.context_map:
            raise CommandError("Use either --context-resource-id or --context-map, not both.")

        workbook_path = Path(options["file"])
        if not workbook_path.is_file():
            raise CommandError(f"Workbook not found: {workbook_path}")

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        sheet = workbook[options["sheet"]] if options.get("sheet") else workbook.active
        rows = self._read_rows(sheet)
        self.stdout.write(f"Simplified Pottery import [{'APPLY' if self.apply else 'DRY-RUN'}]")
        self.stdout.write(f"Workbook: {workbook_path.name}; sheet: {sheet.title}; rows: {len(rows)}")

        totals = Counter()
        for row_number, row in rows:
            try:
                if self.apply:
                    with transaction.atomic():
                        report = self._validate_row(row_number, row)
                        resource_id = self._create_resource(row, report)
                    self.stdout.write(f"  row {row_number}: created {resource_id}")
                else:
                    report = self._validate_row(row_number, row)
                    self.stdout.write(
                        f"  row {row_number}: would create "
                        f"{'one Context and ' if report['totals']['contexts_created'] else ''}"
                        f"one collection, "
                        f"{report['totals']['fragments']} fragment(s)"
                    )
                totals.update(report["totals"])
                totals["rows_ok"] += 1
            except SkipRow as reason:
                totals["rows_skipped"] += 1
                self.stdout.write(f"  row {row_number}: skipped ({reason})")
            except CommandError as error:
                totals["rows_error"] += 1
                self.stderr.write(self.style.ERROR(f"  row {row_number}: {error}"))

        self.stdout.write("Summary:")
        for key in (
            "rows_ok",
            "rows_skipped",
            "rows_error",
            "collections",
            "contexts_created",
            "fragments",
            "unknown_find_values",
            "unknown_chronology_values",
        ):
            self.stdout.write(f"  {key}: {totals[key]}")
        if not self.apply:
            self.stdout.write("Dry-run only. Run again with --apply after reviewing the result.")

    @staticmethod
    def _is_marked_red(cell):
        """Return whether the source cell carries the pink 'remove' marker."""
        color = cell.fill.fgColor
        return color.type == "rgb" and (color.rgb or "").upper().endswith("FFB6C1")

    @classmethod
    def _read_rows(cls, sheet):
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False))
        headers = [clean_cell(cell.value) for cell in header_cells]
        required = {"Form_ID", "Context"}
        missing = required - set(headers)
        if missing:
            raise CommandError(f"Missing required column(s): {', '.join(sorted(missing))}")

        rows = []
        context_index = headers.index("Context")
        for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            values = [cell.value for cell in cells]
            row = dict(zip(headers, values))
            row["_skip_red_context"] = cls._is_marked_red(cells[context_index])
            if any(clean_cell(value) for value in values):
                rows.append((row_number, row))
        return rows

    def _load_context_map(self, context_map_path):
        if not context_map_path:
            return {}
        path = Path(context_map_path)
        if not path.is_file():
            raise CommandError(f"Context map not found: {path}")
        try:
            mapping = json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as error:
            raise CommandError(f"Invalid context map JSON: {error}") from error
        if not isinstance(mapping, dict):
            raise CommandError("Context map must be a JSON object: {\"Excel Context\": \"Arches UUID\"}.")
        return {clean_cell(key): clean_cell(value) for key, value in mapping.items()}

    def _context_for_row(self, row):
        raw_context = clean_cell(row.get("Context"))
        excel_context = normalize_source_context(raw_context)
        if row.get("_skip_red_context"):
            raise SkipRow(f"Context {excel_context!r} is marked red")
        if excel_context.upper().startswith(IGNORED_CONTEXT_PREFIXES):
            raise SkipRow(f"Context {excel_context!r} is excluded")

        resource_id = (
            (self.context_map.get(raw_context) or self.context_map.get(excel_context))
            if self.context_map else self.context_resource_id
        )
        if resource_id:
            return self._get_context_resource(resource_id)

        # Older imports used identifiers such as ``...-C1029``. The current
        # ceramics sheets use ``...-1029`` instead. In both formats the
        # terminal number is the (O) Context Number.
        match = re.search(r"-(?:C)?(\d+)$", excel_context.replace("_", "-"), flags=re.IGNORECASE)
        if not match:
            raise SkipRow(f"cannot derive Context Number from {excel_context!r}")

        context_number = int(match.group(1))
        context_ids = list(
            TileModel.objects.filter(
                resourceinstance__graph_id=CONTEXT_GRAPH_ID,
                data__contains={CONTEXT_NUMBER_NODE_ID: context_number},
            )
            .values_list("resourceinstance_id", flat=True)
            .distinct()
        )
        trench_name = self._source_trench_name(excel_context)
        matching_ids = self._context_ids_in_trench(context_ids, trench_name)
        if len(matching_ids) == 1:
            return Resource.objects.get(resourceinstanceid=matching_ids[0])
        if len(matching_ids) > 1:
            raise CommandError(
                f"more than one local Context {context_number} belongs to {trench_name}"
            )
        if not self.create_missing_contexts:
            if context_ids:
                raise SkipRow(f"no local Context {context_number} belongs to its Trench")
            raise SkipRow(f"no local Context with Context Number {context_number}")
        return self._create_missing_context(excel_context, context_number)

    @staticmethod
    def _source_trench_name(excel_context):
        match = re.match(
            r"^(PAP|MAL)-(TT|T)-([A-Z0-9]+)-(?:C)?\d+$",
            excel_context.upper().replace("_", "-"),
        )
        if not match:
            raise SkipRow(
                f"cannot derive Trench from Context {excel_context!r}"
            )
        # PAP_TT_X and MAL_TT_X are distinct Trench resources. Do not
        # collapse the excavation-area prefix: matching just TT_X could join
        # a source row to a Context from the other area.
        return f"{match.group(1)}_{match.group(2)}_{match.group(3)}"

    def _trench_resources(self):
        if self.trench_resource_ids is not None:
            return self.trench_resource_ids
        trenches = {}
        tiles = (
            TileModel.objects.filter(
                resourceinstance__graph_id=TRENCH_GRAPH_ID,
                data__has_key=TRENCH_NAME_NODE_ID,
            )
            .values_list("resourceinstance_id", "data")
        )
        for resource_id, data in tiles:
            trench_name = clean_cell((data or {}).get(TRENCH_NAME_NODE_ID)).upper()
            if not trench_name:
                continue
            if trench_name in trenches and trenches[trench_name] != str(resource_id):
                raise CommandError(f"more than one local Trench named {trench_name!r}")
            trenches[trench_name] = str(resource_id)
        self.trench_resource_ids = trenches
        return trenches

    def _context_ids_in_trench(self, context_ids, trench_name):
        trench_resource_id = self._trench_resources().get(trench_name)
        if not trench_resource_id:
            return []
        matches = []
        tiles = (
            TileModel.objects.filter(
                resourceinstance_id__in=context_ids,
                data__has_key=CONTEXT_TRENCH_NODE_ID,
            )
            .values_list("resourceinstance_id", "data")
        )
        for resource_id, data in tiles:
            relations = (data or {}).get(CONTEXT_TRENCH_NODE_ID) or []
            if any(
                isinstance(relation, dict)
                and str(relation.get("resourceId")) == trench_resource_id
                for relation in relations
            ):
                matches.append(resource_id)
        return matches

    def _create_missing_context(self, excel_context, context_number):
        trench_name = self._source_trench_name(excel_context)
        cache_key = (trench_name, context_number)
        if cache_key in self.created_contexts:
            context = self.created_contexts[cache_key]
            context._created_by_pottery_import = False
            return context
        trench_resource_id = self._trench_resources().get(trench_name)
        if not trench_resource_id:
            raise SkipRow(f"no local Trench named {trench_name!r}")
        if self.apply:
            context = Resource.objects.create(graph_id=CONTEXT_GRAPH_ID)
            number_node = Node.objects.get(nodeid=CONTEXT_NUMBER_NODE_ID)
            trench_node = Node.objects.get(nodeid=CONTEXT_TRENCH_NODE_ID)
            self._save_tile(
                context, str(number_node.nodegroup_id),
                {CONTEXT_NUMBER_NODE_ID: context_number},
            )
            self._save_tile(
                context, str(trench_node.nodegroup_id),
                {CONTEXT_TRENCH_NODE_ID: [{
                    "resourceId": trench_resource_id,
                    "ontologyProperty": "",
                    "inverseOntologyProperty": "",
                    "resourceXresourceId": str(uuid4()),
                }]},
            )
        else:
            context = Resource(graph_id=CONTEXT_GRAPH_ID)
        context._created_by_pottery_import = True
        self.created_contexts[cache_key] = context
        return context


    @staticmethod
    def _get_context_resource(resource_id):
        try:
            context = Resource.objects.get(resourceinstanceid=UUID(resource_id))
        except (Resource.DoesNotExist, ValueError) as error:
            raise CommandError(f"Context resource does not exist: {resource_id}") from error
        if str(context.graph_id) != CONTEXT_GRAPH_ID:
            raise CommandError(f"Resource {resource_id} is not an (O) Context resource.")
        return context

    @staticmethod
    def _pac_value_id(qid):
        values = Value.objects.filter(
            concept__legacyoid=f"{PAC_ENTITY_URL}{qid}",
            valuetype_id="prefLabel",
        ).order_by("language_id")
        english = values.filter(language_id="en").first()
        value = english or values.first()
        if value is None:
            raise CommandError(f"No local Arches value found for PAC {qid}.")
        return str(value.valueid)

    @staticmethod
    def _collection_exists(form_id):
        """Keep repeated --apply runs from creating duplicate collections."""
        return TileModel.objects.filter(
            resourceinstance__graph_id=POTTERY_GRAPH_ID,
            data__contains={FORM_ID_NODE_ID: form_id},
        ).exists()

    @staticmethod
    def _find_values(raw_value, unknown_values):
        resolved = []
        for label in (clean_cell(value) for value in clean_cell(raw_value).split(",")):
            if not label:
                continue
            value_id = resolve_dictionary_value(FINDS_DICTIONARY_ID, label)
            if value_id:
                resolved.append(value_id)
            else:
                unknown_values.append(label)
        return list(dict.fromkeys(resolved))

    @staticmethod
    def _resolve_period_label(label):
        """Resolve one complete PAC period label, including a transition."""
        value_id = resolve_dictionary_value(POTTERY_DICTIONARY_CHRONOLOGY, label)
        if value_id:
            return value_id

        # Accept spreadsheet spelling such as "Late Hellenistic/Early Roman"
        # for PAC's "Late Hellenistic / Early Roman".
        normalized_slash_label = re.sub(r"\s*/\s*", " / ", label)
        return resolve_dictionary_value(
            POTTERY_DICTIONARY_CHRONOLOGY,
            normalized_slash_label,
        )

    @staticmethod
    def _range_period_values(label):
        """Split a source range into locally available chronology periods.

        A slash is retained because it can form a PAC transition period. A
        hyphen and parentheses are range separators. Spaces can also separate
        consecutive period labels, for example ``Early Hellenistic Middle
        Hellenistic``. The longest matching dictionary label always wins.
        """
        dictionary = get_dictionary_index(POTTERY_DICTIONARY_CHRONOLOGY)
        remaining = normalize_dictionary_label(label)
        remaining = re.sub(r"\s*[-–—]\s*", " ", remaining)
        remaining = re.sub(r"[()]", " ", remaining)
        remaining = re.sub(r"\s*/\s*", " / ", remaining)
        remaining = re.sub(r"\s+", " ", remaining).strip()
        candidates = sorted(
            dictionary.values_by_label.items(),
            key=lambda item: len(item[0]),
            reverse=True,
        )
        resolved = []

        while remaining:
            match = next(
                (
                    (period_label, value_id)
                    for period_label, value_id in candidates
                    if remaining == period_label
                    or remaining.startswith(f"{period_label} ")
                    or remaining.startswith(f"{period_label} /")
                ),
                None,
            )
            if not match:
                return resolved, remaining

            period_label, value_id = match
            resolved.append(value_id)
            remaining = remaining[len(period_label):].strip()
            if remaining.startswith("/"):
                remaining = remaining[1:].strip()

        return resolved, ""
    @staticmethod
    def _century_suffix(century):
        if 10 <= century % 100 <= 20:
            return "th"
        return {1: "st", 2: "nd", 3: "rd"}.get(century % 10, "th")

    @staticmethod
    def _expand_chronology_abbreviations(value):
        return re.sub(
            r"\b(LCL|LH|EH|MH|ER)\b",
            lambda match: CHRONOLOGY_ABBREVIATIONS[match.group(1).upper()],
            clean_cell(value),
            flags=re.IGNORECASE,
        )

    @classmethod
    def _century_range_labels(cls, label):
        """Return every canonical century covered by a source century range."""
        match = re.fullmatch(
            r"\s*(\d+)(?:st|nd|rd|th)?\s*(?:c(?:entury)?\.?)?\s*[-–—]\s*"
            r"(\d+)(?:st|nd|rd|th)?\s*(?:c(?:entury)?\.?)?\s*(AD|CE|BC|BCE)\s*",
            clean_cell(label),
            flags=re.IGNORECASE,
        )
        if not match:
            return []
        start, end = (int(value) for value in match.group(1, 2))
        if start > end:
            return []
        # Only low-numbered source ranges denote centuries. Higher values
        # in the amphora sheets are calendar years, even when an earlier
        # conversion added ordinal or ``c.`` markers.
        if max(start, end) > 10:
            return []
        era = {"AD": "CE", "CE": "CE", "BC": "BCE", "BCE": "BCE"}[match.group(3).upper()]
        return [
            f"{century}{cls._century_suffix(century)} c. {era}"
            for century in range(start, end + 1)
        ]

    @classmethod
    def _period_values(cls, raw_value, unknown_values):
        """Resolve one or more Excel periods separated with a vertical bar.

        A question mark is an import marker for the Dating > Uncertain field;
        it is not part of the PAC chronology label used for lookup.
        """
        resolved = []
        for label in (
            clean_cell(value)
            for value in re.split(
                r"\s*(?:\||,)\s*", cls._expand_chronology_abbreviations(raw_value)
            )
        ):
            label = clean_cell(label.replace("?", ""))
            if not label:
                continue
            value_id = cls._resolve_period_label(label)
            if value_id:
                resolved.append(value_id)
                continue

            century_labels = cls._century_range_labels(label)
            if century_labels:
                for century_label in century_labels:
                    value_id = cls._resolve_period_label(century_label)
                    if value_id:
                        resolved.append(value_id)
                    else:
                        unknown_values.append(century_label)
                continue

            range_value_ids, unknown_tail = cls._range_period_values(label)
            resolved.extend(range_value_ids)
            if unknown_tail:
                unknown_values.append(unknown_tail)
        return list(dict.fromkeys(resolved))

    @staticmethod
    def _is_uncertain_period(raw_value):
        """Return whether the source Dating value is explicitly uncertain."""
        return "?" in clean_cell(raw_value)

    def _category_data(self, row, prefix):
        prefix_key = f"{prefix}_"
        raw_type = clean_cell(row.get(f"{prefix_key}Pottery_Type"))
        diagnostic = cell_number(row.get(f"{prefix_key}Number_of_Diagnostic_Fragments"))
        undiagnostic = cell_number(row.get(f"{prefix_key}Number_of_Undiagnostic_Fragments"))
        period = clean_cell(row.get(f"{prefix_key}General_Chronology_of_Category_Period"))
        comment = clean_cell(row.get(f"{prefix_key}General_Chronology_of_Category_Comment"))
        category_remarks = clean_cell(row.get(f"{prefix_key}Remarks_from_Pottery_Category_Form"))
        no_material = to_boolean(row.get(f"{prefix_key}No_Material"))
        special_finds = to_boolean(row.get(f"{prefix_key}Contains_Special_Finds"))
        if (
            not raw_type
            and diagnostic in (None, 0)
            and undiagnostic in (None, 0)
            and not no_material
            and not special_finds
            and not period
            and not comment
            and not category_remarks
        ):
            return None

        return {
            "type_value": self._pac_value_id(CATEGORY_TYPES[prefix]),
            "diagnostic": diagnostic,
            "undiagnostic": undiagnostic,
            "no_material": no_material,
            "special_finds": special_finds,
            "period": period,
            "comment": comment,
            "category_remarks": category_remarks,
            "finds_remarks": clean_cell(row.get(f"{prefix_key}Remarks_from_Finds_Recording_Form")),
        }

    def _validate_row(self, row_number, row):
        form_id = clean_cell(row.get("Form_ID"))
        if not form_id:
            raise CommandError("Form_ID is empty.")
        if self._collection_exists(form_id):
            raise SkipRow(f"Pottery Collection with Form_ID {form_id!r} already exists")
        context = self._context_for_row(row)
        unknown_values = []
        archaeological_remains = self._find_values(row.get("Archaeological_Remains"), unknown_values)
        special_finds = self._find_values(row.get("Special_Finds"), unknown_values)
        unknown_periods = []
        context_period_value_ids = self._period_values(
            row.get("General_Context_Chronology_Period"),
            unknown_periods,
        )
        context_period_uncertain = self._is_uncertain_period(
            row.get("General_Context_Chronology_Period")
        )

        categories = []
        for prefix in CATEGORY_TYPES:
            category = self._category_data(row, prefix)
            if category is None:
                continue
            if category["period"]:
                category["period_value_ids"] = self._period_values(
                    category["period"],
                    unknown_periods,
                )
            category["period_uncertain"] = self._is_uncertain_period(
                category["period"]
            )
            categories.append(category)

        for value in unknown_values:
            self.stderr.write(self.style.WARNING(f"  row {row_number}: ignored unknown find value {value!r}"))
        for value in unknown_periods:
            self.stderr.write(self.style.WARNING(f"  row {row_number}: ignored unknown chronology {value!r}"))

        return {
            "context": context,
            "archaeological_remains": archaeological_remains,
            "special_finds": special_finds,
            "context_period_value_ids": context_period_value_ids,
            "context_period_uncertain": context_period_uncertain,
            "categories": categories,
            "totals": Counter(
                collections=1,
                contexts_created=int(getattr(context, "_created_by_pottery_import", False)),
                fragments=len(categories),
                unknown_find_values=len(unknown_values),
                unknown_chronology_values=len(unknown_periods),
            ),
        }

    @staticmethod
    def _save_tile(resource, nodegroup_id, data, parent_tile=None):
        tile = Tile.get_blank_tile_from_nodegroup_id(
            nodegroup_id,
            resourceid=str(resource.resourceinstanceid),
        )
        tile.parenttile = parent_tile
        tile.data.update(data)
        tile.save()
        return tile

    def _create_resource(self, row, report):
        resource = Resource.objects.create(graph_id=POTTERY_GRAPH_ID)
        context_data = {
            CONTEXT_NODE_ID: [{
                "resourceId": str(report["context"].resourceinstanceid),
                "ontologyProperty": "",
                "inverseOntologyProperty": "",
                "resourceXresourceId": str(uuid4()),
            }],
        }
        if report["archaeological_remains"]:
            context_data[ARCHAEOLOGICAL_REMAINS_NODE_ID] = report["archaeological_remains"]
        if report["special_finds"]:
            context_data[SPECIAL_FINDS_NODE_ID] = report["special_finds"]
        if clean_cell(row.get("Remarks_about_Context")):
            context_data[CONTEXT_REMARKS_NODE_ID] = localized_string(
                clean_cell(row["Remarks_about_Context"])
            )
        context_tile = self._save_tile(resource, CONTEXT_NODEGROUP_ID, context_data)

        self._save_tile(resource, FORM_ID_NODEGROUP_ID, {FORM_ID_NODE_ID: clean_cell(row["Form_ID"])})
        last_shred = cell_number(row.get("Last_Shred_No"))
        if last_shred is not None:
            self._save_tile(resource, LAST_SHRED_NODEGROUP_ID, {LAST_SHRED_NODE_ID: last_shred})

        chronology_data = {}
        if report["context_period_value_ids"]:
            chronology_data[CONTEXT_PERIOD_NODE_ID] = report["context_period_value_ids"]
        if clean_cell(row.get("General_Context_Chronology_Comment")):
            chronology_data[CONTEXT_CHRONOLOGY_COMMENT_NODE_ID] = localized_string(
                clean_cell(row["General_Context_Chronology_Comment"])
            )
        if chronology_data:
            # Uncertain is required by the current Dating card. Explicitly
            # write False when the Excel value has no question-mark marker.
            chronology_data[CONTEXT_UNCERTAIN_NODE_ID] = report["context_period_uncertain"]
            self._save_tile(
                resource,
                CONTEXT_CHRONOLOGY_NODEGROUP_ID,
                chronology_data,
                parent_tile=context_tile,
            )

        for category in report["categories"]:
            fragment_data = {
                POTTERY_TYPE_NODE_ID: category["type_value"],
                NO_MATERIAL_NODE_ID: category["no_material"],
                CONTAINS_SPECIAL_FINDS_NODE_ID: category["special_finds"],
            }
            if category["diagnostic"] is not None:
                fragment_data[DIAGNOSTIC_NODE_ID] = category["diagnostic"]
            if category["undiagnostic"] is not None:
                fragment_data[UNDIAGNOSTIC_NODE_ID] = category["undiagnostic"]
            if category["category_remarks"]:
                fragment_data[CATEGORY_REMARKS_NODE_ID] = localized_string(
                    category["category_remarks"]
                )
            if category["finds_remarks"]:
                fragment_data[FINDS_REMARKS_NODE_ID] = localized_string(
                    category["finds_remarks"]
                )
            fragment_tile = self._save_tile(resource, FRAGMENT_NODEGROUP_ID, fragment_data)

            chronology_data = {}
            if category.get("period_value_ids"):
                chronology_data[CATEGORY_PERIOD_NODE_ID] = category["period_value_ids"]
            if category["comment"]:
                chronology_data[CATEGORY_CHRONOLOGY_COMMENT_NODE_ID] = localized_string(
                    category["comment"]
                )
            if chronology_data:
                chronology_data[CATEGORY_UNCERTAIN_NODE_ID] = category["period_uncertain"]
                self._save_tile(
                    resource,
                    CATEGORY_CHRONOLOGY_NODEGROUP_ID,
                    chronology_data,
                    parent_tile=fragment_tile,
                )

        return resource.resourceinstanceid
