"""Import detailed Pottery Records from A, TW, PW, SV and L Excel sheets.

The detailed sheets are linked by their full ``Pottery Collection`` source
identifier, e.g. ``PAP-TT-X-912``.  They are deliberately not linked merely
by Form ID: the same Context Number can occur in both PAP and MAL trenches.
"""

from collections import Counter
from pathlib import Path
import re
from uuid import uuid4

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import Workbook, load_workbook

from arches.app.models.models import Resource, TileModel
from arches.app.models.tile import Tile
from arches.app.utils.date_utils import ExtendedDateFormat
from arches_for_excavation.management.commands.import_collection_excel import (
    CONTEXT_NODE_ID,
    POTTERY_GRAPH_ID,
    Command as CollectionImportCommand,
    SkipRow,
    normalize_source_context,
)
from arches_for_excavation.utils.pottery.common import (
    clean_cell,
    ensure_tile_parent,
    localized_string,
    to_boolean,
)
from arches_for_excavation.utils.pottery.concept_lookup import (
    apply_dictionary_alias,
    format_concept_tile_value,
    get_node_datatype,
    resolve_dictionary_value,
)
from arches_for_excavation.utils.pottery.generic_record_parser import normalize_partial_century
from arches_for_excavation.utils.pottery.constants import (
    POTTERY_DICTIONARY_FORM,
    POTTERY_DICTIONARY_STATE_OF_PRESERVATION,
    POTTERY_DICTIONARY_SUB_CATEGORY,
    POTTERY_DICTIONARY_SURFACE_TREATMENT,
    POTTERY_RECORD_TYPES,
)


CATEGORY_RECORD_TYPES = {
    "A": "amphorae",
    "TW": "table-ware",
    "PW": "plain-ware",
    "SV": "storage-vessel",
    "L": "lamp",
}

FREE_TEXT_TYPE_RECORD_TYPES = {"plain-ware", "storage-vessel"}
FILENAME_CATEGORY = re.compile(r"(?:^|[_-])(A|TW|SV|PW|L)(?:[_-]|$)", re.IGNORECASE)

# ``flat hole`` is a known no-comma form found in the source sheets for two
# morphology values. Keep this correction narrow: spaces normally belong to
# one dictionary label and must not be treated as separators.
MULTIPLE_MORPHOLOGY_NORMALIZATIONS = {
    "flat hole": "flat, hole",
    "everted flat": "everted, flat",
}

MULTIPLE_VESSEL_FORM_NORMALIZATIONS = {
    "amphora jug": "Amphora, Jug",
    "jug amphora": "Jug, Amphora",
}

MULTIPLE_SUBCATEGORY_NORMALIZATIONS = {
    "colour-coated ware black gloss ware": "Colour-Coated Ware, Black Gloss Ware",
}

VESSEL_PART_NORMALIZATIONS = {
    "up part aa": "upper part",
    "upper part rn": "upper part",
}

# Preserve the source wording for partial-century dates in Comment while
# storing the requested exact EDTF boundaries. The two-century form is a
# multi-select chronology, not a date range.
SPECIAL_CHRONOLOGY_NORMALIZATIONS = {
    "1st 2nd c ce": {
        "periods": "1st c. CE|2nd c. CE",
    },
}

# Nodes which are present in all five record graphs but are not part of the
# original web-workflow parser configuration.
EXTRA_NODES = {
    "amphorae": {
        "mgNo": "08aab7c3-edf8-48f4-b601-2fd660e7d6f1",
        "author": "f9e6e252-923a-4eed-b69c-d0ad00f8dc1f",
        "vesselForm": "eeb1eb0e-f650-4f6c-a75a-291483c47db8",
        "subcategory": "0fcf663f-7491-437c-b1bf-f62390434117",
        "stateOfPreservation": "74c65f24-af0a-4b04-8678-a88be7102beb",
        "surfaceTreatment": "b19dcb7b-29a4-4bba-ada4-e45328549c3c",
        "specialFindId": "a52e4cf8-17fe-4359-8167-4f2a0728d4f2",
    },
    "table-ware": {
        "mgNo": "e7a5e00c-a5fa-41c9-97bd-9ecea21760ce",
        "author": "1bc1eb28-451a-46a7-881f-79c56e197d31",
        "vesselForm": "38c60f78-015a-4a20-99a4-58331b399b23",
        "subcategory": "e1cc37a7-7be8-4cd2-9f3f-8ca3f84d1769",
        "stateOfPreservation": "822a6fdb-3673-456a-8c16-f05ee69641fc",
        "surfaceTreatment": "b16d1ad7-b37f-42b6-b525-3bee02a4800c",
        "specialFindId": "83510c26-35b4-44be-b13a-e7bd93a717ca",
    },
    "plain-ware": {
        "mgNo": "4ba8d85f-e410-40a1-a444-9db0feaff10c",
        "author": "2b4823f7-3503-4f5e-bbd6-8057acab8651",
        "vesselForm": "9662494d-2679-4f70-9b7f-e4c2f513ee05",
        "subcategory": "9563d7a9-a6f3-4ddd-9623-78cc3bfc6890",
        "stateOfPreservation": "81ca8aa4-24c8-484b-97eb-75dd407e163c",
        "surfaceTreatment": "65837d73-e414-43bb-bf30-55601750f7d6",
        "specialFindId": "6e2ff31b-d21c-443e-918c-a9363c2f8828",
    },
    "storage-vessel": {
        "mgNo": "1f8927a3-7432-4d00-8436-11b96e76a843",
        "author": "b2d01d43-82e3-4bb7-b2bf-f79d37ad2b70",
        "vesselForm": "ecd44720-8cd7-438e-9c97-5b7886960fdd",
        "subcategory": "9ae41f28-d43c-4569-9b27-625a5cc43085",
        "stateOfPreservation": "d37e518b-59f5-402a-9dcd-1f350a6439dc",
        "surfaceTreatment": "3498af89-5372-43dc-8abb-843ff3c2fe14",
        "specialFindId": "cc79936a-87c4-432f-a16c-ab0442720837",
    },
    "lamp": {
        "mgNo": "5694d5a1-8501-43dd-9c55-fb1b53d50396",
        "author": "cf1967f5-0018-44d5-a733-95afe35e5a65",
        "vesselForm": "8176d8fe-a691-46fa-bebe-efdae625e929",
        "subcategory": "3226a7a9-d719-42df-9b27-33bd7bd667dd",
        "stateOfPreservation": "14de5b24-b037-4b9f-8ed5-3b58fac42737",
        "surfaceTreatment": "5638eb7a-8ebf-4156-8f9c-90ebbece67ee",
        "specialFindId": "0089a6b8-e360-4a91-ab3b-7d2a5a627f06",
    },
}

FIELD_HEADERS = {
    "formNo": ("Form ID",),
    "context": ("Pottery Collection", "Context"),
    "pNo": ("P number", "P Number", "L number", "L Number"),
    "count": ("Quantity",),
    "vesselPart": ("Vessel Part",),
    "type": ("Type",),
    "typeUncertain": ("Type Uncertainty",),
    "morphology": ("Morphology",),
    "chronology": ("Period",),
    "chronologyUncertain": ("Uncertain",),
    "provenance": ("Provenance",),
    "provenanceUncertain": ("Provenance Uncertainity", "Provenance Uncertainty"),
    "drawn": ("Drawing",),
    "photo": ("Photo",),
    "comment": ("Comment",),
    "mgNo": ("MG no", "TW MG no", "PW MG no", "PW/SV MG no", "Lamp MG no"),
    "author": ("Author",),
    "vesselForm": ("Form", "Vessel Form"),
    "subcategory": ("Sub-Category", "Subcategory"),
    "surfaceTreatment": ("Surface Treatment",),
    "stateOfPreservation": ("State of Preservation", "State Of Preservation"),
    "specialFindId": ("Special Find ID",),
}


class Command(BaseCommand):
    help = "Import detailed A/TW/PW/SV/L Pottery Records linked by full Pottery Collection Context."

    def add_arguments(self, parser):
        source = parser.add_mutually_exclusive_group(required=True)
        source.add_argument("--directory", help="Directory containing detailed pottery .xlsx files.")
        source.add_argument("--file", action="append", dest="files", help="One detailed .xlsx file; may be repeated.")
        parser.add_argument("--apply", action="store_true", help="Create records. Without this flag the command is a dry-run.")
        parser.add_argument("--dry-run", action="store_true", help="Explicitly validate only; records are never created.")
        parser.add_argument("--dataset", default="pottery-records", help="Stable prefix used for idempotent source row IDs.")
        parser.add_argument("--missing-collections-report", help="Optional .xlsx path for the dry-run list of Contexts without a usable Pottery Collection.")
        parser.add_argument(
            "--audit-contexts",
            action="store_true",
            help="Do not write records; report every Context or collection-link problem to an .xlsx file.",
        )
        parser.add_argument("--context-issues-report", help="Optional .xlsx path for --audit-contexts output.")
        parser.add_argument(
            "--audit-unknown-concepts",
            action="store_true",
            help="Do not write records; report every unresolved dictionary value to an .xlsx file.",
        )
        parser.add_argument(
            "--unknown-concepts-report",
            help="Optional .xlsx path for --audit-unknown-concepts output.",
        )
        parser.add_argument(
            "--audit-type-values", action="store_true",
            help="Report unique raw Type values in Plain Ware and Storage Vessel sheets without importing records.",
        )
        parser.add_argument("--type-values-report", help="Optional .xlsx path for --audit-type-values output.")

    def handle(self, *args, **options):
        if options["apply"] and (options["dry_run"] or options["audit_unknown_concepts"] or options["audit_contexts"] or options["audit_type_values"]):
            raise CommandError("--apply cannot be used with --dry-run or an audit option.")
        self.audit_unknown_concepts = options["audit_unknown_concepts"]
        self.audit_contexts = options["audit_contexts"]
        self.audit_type_values = options["audit_type_values"]
        self.unknown_concept_values = {}
        self.type_values = {}
        self.apply = options["apply"]
        self.dataset = clean_cell(options["dataset"]) or "pottery-records"
        self.context_importer = CollectionImportCommand()
        self.context_importer.apply = False
        self.context_importer.create_missing_contexts = False
        self.context_importer.context_resource_id = None
        self.context_importer.context_map = {}
        self.context_importer.created_contexts = {}
        self.context_importer.trench_resource_ids = None
        self.collection_ids_by_context = {}
        self.missing_collections = {}

        files = self._source_files(options)
        if not files:
            raise CommandError("No detailed A/TW/PW/SV/L .xlsx files found.")
        self.unknown_concepts_report_path = self._unknown_concepts_report_path(options, files)
        self.missing_collections_report_path = self._context_issues_report_path(options, files)
        self.type_values_report_path = self._type_values_report_path(options, files)

        totals = Counter()
        self.stdout.write(f"Pottery records import [{'APPLY' if self.apply else 'DRY-RUN'}]")
        self.stdout.write(f"Workbooks: {len(files)}")
        for path in files:
            record_type = self._record_type_from_path(path)
            if not record_type:
                totals["workbooks_skipped"] += 1
                self.stderr.write(f"Skipped {path.name}: cannot derive pottery category from filename")
                continue
            totals.update(self._process_workbook(path, record_type))

        self.stdout.write("Summary:")
        for key in (
            "workbooks_recognized", "workbooks_skipped", "rows_seen", "created",
            "would_create", "unchanged", "skipped_empty", "skipped_red_context",
            "missing_collections", "duplicate_collections", "invalid_contexts",
            "invalid_quantity", "unknown_concepts", "unknown_special_finds", "valid_contexts", "type_value_rows", "errors",
        ):
            self.stdout.write(f"  {key}: {totals[key]}")
        if not self.apply:
            if self.audit_type_values:
                self._write_type_values_report()
                self.stdout.write(f"Unique PW/SV Type values: {len(self.type_values)}")
                self.stdout.write("Type-values audit only. No records were created.")
            else:
                if not self.audit_contexts:
                    self._write_missing_collections()
                self._write_missing_collection_report()
                if self.audit_contexts:
                    self.stdout.write("Context audit only. No records were created.")
                elif self.audit_unknown_concepts:
                    self._write_unknown_concepts_report()
                    self._write_unknown_concepts()
                    self.stdout.write("Unknown-concepts audit only. No records were created.")
                else:
                    self.stdout.write("Dry-run only. Add --apply to create records.")

    @staticmethod
    def _is_marked_red(cell):
        fill = getattr(cell, "fill", None)
        color = getattr(fill, "fgColor", None)
        if color is None:
            return False
        return color.type == "rgb" and (color.rgb or "").upper().endswith("FFB6C1")

    @staticmethod
    def _record_type_from_path(path):
        stem = path.stem.casefold()
        # Historic Polish source files are often named ``amfory`` or
        # ``amfory2`` rather than using the category token ``A``.
        if re.search(r"(?:^|[_-])amfor(?:y|a)\d*(?:[_-]|$)", stem):
            return "amphorae"
        match = FILENAME_CATEGORY.search(path.stem)
        return CATEGORY_RECORD_TYPES.get(match.group(1).upper()) if match else ""

    @classmethod
    def _source_files(cls, options):
        if options.get("directory"):
            directory = Path(options["directory"])
            if not directory.is_dir():
                raise CommandError(f"Directory not found: {directory}")
            return [
                path for path in sorted(directory.rglob("*.xlsx"))
                if not path.name.startswith("~$")
                and "audyt" not in {part.casefold() for part in path.parts}
                and "_m_" not in path.stem.casefold()
                and "main" not in path.stem.casefold()
                and cls._record_type_from_path(path)
            ]
        paths = [Path(value) for value in options.get("files", [])]
        missing = [str(path) for path in paths if not path.is_file()]
        if missing:
            raise CommandError(f"Workbook not found: {', '.join(missing)}")
        return paths

    def _process_workbook(self, path, record_type):
        totals = Counter(workbooks_recognized=1)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows = sheet.iter_rows(values_only=False)
                try:
                    header_cells = next(rows)
                except StopIteration:
                    continue
                headers = [clean_cell(cell.value) for cell in header_cells]
                context_header = next(
                    (header for header in FIELD_HEADERS["context"] if header in headers),
                    None,
                )
                if context_header is None:
                    continue
                context_index = headers.index(context_header)
                self.stdout.write(f"Workbook: {path.name}; sheet: {sheet.title} ({record_type})")
                for row_number, cells in enumerate(rows, start=2):
                    values = [cell.value for cell in cells]
                    row = dict(zip(headers, values))
                    if self.audit_type_values:
                        if record_type in {"plain-ware", "storage-vessel"}:
                            raw_type = self._value(row, "type")
                            if raw_type:
                                source = f"{path.name}:{sheet.title}:{row_number}"
                                self._record_type_value(record_type, raw_type, source)
                                totals["type_value_rows"] += 1
                        continue
                    if self._is_marked_red(cells[context_index]):
                        source = f"{path.name}:{sheet.title}:{row_number}"
                        context_label = normalize_source_context(self._value(row, "context")) or "<empty>"
                        self._record_missing_collection(context_label, record_type, source, "marked red (excluded)")
                        totals["skipped_red_context"] += 1
                        self.stdout.write(f"  {source}: skipped Pottery Collection marked red")
                        continue
                    totals.update(self._process_row(path, sheet.title, row_number, row, record_type))
        finally:
            workbook.close()
        return totals

    def _process_row(self, path, sheet_name, row_number, row, record_type):
        totals = Counter(rows_seen=1)
        form_no = self._value(row, "formNo")
        context_label = normalize_source_context(self._value(row, "context"))
        source = f"{path.name}:{sheet_name}:{row_number}"
        if not context_label and not any(clean_cell(value) for value in row.values()):
            totals["skipped_empty"] += 1
            return totals
        # Form ID is descriptive data, not the key used to find a collection.
        # A blank Excel cell must remain blank in Arches.
        if not context_label:
            totals["errors"] += 1
            self._record_missing_collection(context_label or "<empty>", record_type, source, "missing Pottery Collection")
            self.stderr.write(f"  {source}: missing Pottery Collection")
            return totals

        if not self._has_record_data(row):
            totals["skipped_empty"] += 1
            return totals

        if self.audit_contexts:
            collection_ids, context_error = self._collection_ids(context_label)
            if context_error:
                totals["invalid_contexts"] += 1
                self._record_missing_collection(context_label, record_type, source, context_error)
                return totals
            if not collection_ids:
                totals["missing_collections"] += 1
                self._record_missing_collection(context_label, record_type, source, "collection not found")
                return totals
            if len(collection_ids) != 1:
                totals["duplicate_collections"] += 1
                self._record_missing_collection(context_label, record_type, source, f"{len(collection_ids)} collections found")
                return totals
            totals["valid_contexts"] += 1
            return totals

        record, preparation_totals = self._prepare_record(row, record_type, source)
        totals.update(preparation_totals)
        collection_ids, context_error = self._collection_ids(context_label)
        if context_error:
            totals["invalid_contexts"] += 1
            self._record_missing_collection(context_label, record_type, source, context_error)
            self.stderr.write(f"  {source}: {context_error}")
            return totals
        collection_ids, collection_created, creation_error = self._create_missing_collection_if_enabled(
            context_label, collection_ids,
        )
        if creation_error:
            totals["invalid_contexts"] += 1
            self._record_missing_collection(context_label, record_type, source, creation_error)
            self.stderr.write(f"  {source}: {creation_error}")
            return totals
        if collection_created:
            totals["collections_created" if self.apply else "collections_would_create"] += 1
            self.stdout.write(
                f"  {source}: "
                f"{'created' if self.apply else 'would create'} Pottery Collection for {context_label!r}"
            )
        if not collection_ids:
            totals["missing_collections"] += 1
            self._record_missing_collection(context_label, record_type, source, "collection not found")
            self.stdout.write(f"  {source}: skipped {context_label!r} (collection not found)")
            return totals
        if len(collection_ids) != 1:
            totals["duplicate_collections"] += 1
            self._record_missing_collection(context_label, record_type, source, f"{len(collection_ids)} collections found")
            self.stderr.write(f"  {source}: Context {context_label!r} has {len(collection_ids)} collections")
            return totals

        legacyid = self._legacyid(record_type, context_label, form_no, record["pNo"], row_number)
        config = POTTERY_RECORD_TYPES[record_type]
        if Resource.objects.filter(graph_id=config["graph_id"], legacyid=legacyid).exists():
            totals["unchanged"] += 1
            self.stdout.write(f"  {source}: unchanged")
            return totals
        if not self.apply:
            totals["would_create"] += 1
            self.stdout.write(f"  {source}: would create {config['label']} Pottery Record")
            return totals

        with transaction.atomic():
            resource = Resource.objects.create(graph_id=config["graph_id"], legacyid=legacyid)
            self._save_record(resource, record_type, record, str(collection_ids[0]))
            resource.name = f"{config['label']} {form_no or record['pNo'] or 'record'} row {row_number}"
            resource.save(update_fields=["name"])
        totals["created"] += 1
        self.stdout.write(f"  {source}: created {resource.resourceinstanceid}")
        return totals

    def _prepare_record(self, row, record_type, source):
        config = POTTERY_RECORD_TYPES[record_type]
        fields = config["fields"]
        extra = EXTRA_NODES[record_type]
        totals = Counter()
        record = {
            "formNo": self._value(row, "formNo"),
            "pNo": self._value(row, "pNo"),
            "count": self._number(self._value(row, "count")),
            "mgNo": self._value(row, "mgNo"),
            "author": self._value(row, "author"),
            "comment": self._value(row, "comment"),
            "chronologyBoundaries": None,
        }
        source_notes = []
        if self._value(row, "count") and record["count"] is None:
            totals["invalid_quantity"] += 1
            self.stderr.write(f"  {source}: invalid Quantity {self._value(row, 'count')!r}; omitted")

        for key in ("drawn", "photo"):
            raw_value = self._value(row, key)
            record[key] = to_boolean(raw_value) if raw_value else None
        for value_key, uncertainty_key in (
            ("type", "typeUncertain"),
            ("chronology", "chronologyUncertain"),
            ("provenance", "provenanceUncertain"),
        ):
            explicit_value = self._value(row, uncertainty_key)
            source_value = self._value(row, value_key)

            record[uncertainty_key] = (
                to_boolean(explicit_value)
                if explicit_value
                else (True if self._has_uncertainty_marker(source_value) else None)
            )
        dictionaries = dict(config.get("dictionary_fields", {}))
        dictionaries.update({
            "vesselForm": POTTERY_DICTIONARY_FORM,
            "subcategory": POTTERY_DICTIONARY_SUB_CATEGORY,
            "surfaceTreatment": POTTERY_DICTIONARY_SURFACE_TREATMENT,
            "stateOfPreservation": POTTERY_DICTIONARY_STATE_OF_PRESERVATION,
        })
        node_ids = dict(fields)
        node_ids.update(extra)
        for key in (
            "vesselPart", "type", "morphology", "provenance", "vesselForm",
            "subcategory", "surfaceTreatment", "stateOfPreservation",
        ):
            source_value = self._value(row, key)
            if key == "type" and record_type in FREE_TEXT_TYPE_RECORD_TYPES:
                record[key] = self._non_uncertainty_value(source_value)
                continue
            if key in {"type", "provenance"}:
                raw_value = self._strip_uncertainty_marker(source_value)[0]
            else:
                raw_value = (
                    source_value
                    if key == "vesselPart"
                    else self._non_uncertainty_value(source_value)
                )
                if self._has_uncertainty_marker(source_value):
                    source_notes.append(f"{FIELD_HEADERS[key][0]}: {source_value}")
            record[key] = self._concept_values(
                key, raw_value, dictionaries[key], node_ids[key], config, source, totals,
            )

        unknown_periods = []
        source_chronology = self._value(row, "chronology")
        chronology_normalization = self._special_chronology(source_chronology)
        chronology_value = source_chronology
        if chronology_normalization:
            chronology_value = chronology_normalization["periods"]
            record["chronologyBoundaries"] = chronology_normalization.get("boundaries")
            if chronology_normalization.get("preserve_in_comment"):
                source_notes.append(f"{FIELD_HEADERS['chronology'][0]}: {source_chronology}")

        record["chronology"] = CollectionImportCommand._period_values(
            chronology_value, unknown_periods,
        )
        for value in unknown_periods:
            tail_normalization = self._special_chronology(value)
            if tail_normalization:
                record["chronology"] = list(dict.fromkeys(record["chronology"] + CollectionImportCommand._period_values(tail_normalization["periods"], []))); record["chronologyBoundaries"] = tail_normalization.get("boundaries"); source_notes.extend([f"{FIELD_HEADERS['chronology'][0]}: {source_chronology}"] if tail_normalization.get("preserve_in_comment") else []); continue
            totals["unknown_concepts"] += 1
            self._record_unknown("chronology", value, source)
        if source_notes:
            comment = record["comment"]
            source_comment = "\n".join(source_notes)
            record["comment"] = "\n".join(
                value for value in (comment, source_comment) if value
            )
        record["specialFindId"] = self._special_find_relations(
            self._value(row, "specialFindId"), source, totals,
        )
        return record, totals

    @staticmethod
    def _calendar_date_range_boundaries(raw_value):
        """Return EDTF-year boundaries for an unqualified numeric date range."""
        raw_value = clean_cell(raw_value)
        era_match = re.search(r"\b(AD|CE|BC|BCE)\s*$", raw_value, flags=re.IGNORECASE)
        if not era_match:
            return None
        years = [int(value) for value in re.findall(r"\d+", raw_value)]
        if not 1 <= len(years) <= 2:
            return None
        start, end = (years[0], years[-1])
        # Small values such as 4-7 AD represent centuries. Larger values are
        # calendar years, including source values incorrectly labelled ``c.``.
        if max(start, end) <= 10:
            return None
        if era_match.group(1).upper() in {"AD", "CE"}:
            return tuple(sorted((start, end)))
        # EDTF uses astronomical year numbering: 1 BCE is year 0.
        return tuple(sorted((1 - start, 1 - end)))

    @classmethod
    def _special_chronology(cls, raw_value):
        normalized = re.sub(r"[.]", "", clean_cell(raw_value))
        normalized = re.sub(r"\s+", " ", normalized).casefold()
        normalized = re.sub(r"\s*[-–—]\s*", " ", normalized)
        special = SPECIAL_CHRONOLOGY_NORMALIZATIONS.get(normalized)
        if special:
            return special
        partial_century = normalize_partial_century(raw_value)
        if partial_century:
            periods, boundaries = partial_century
            return {
                "periods": periods,
                "boundaries": boundaries,
                "preserve_in_comment": True,
            }
        boundaries = cls._calendar_date_range_boundaries(raw_value)
        return {"periods": "", "boundaries": boundaries} if boundaries else None

    def _concept_values(self, key, raw_value, dictionary, node_id, config, source, totals):
        raw_value = clean_cell(raw_value)
        if not raw_value:
            return [] if get_node_datatype(node_id) == "concept-list" else ""
        if key == "morphology":
            raw_value = MULTIPLE_MORPHOLOGY_NORMALIZATIONS.get(
                raw_value.casefold(), raw_value,
            )
        elif key == "vesselForm":
            raw_value = MULTIPLE_VESSEL_FORM_NORMALIZATIONS.get(
                raw_value.casefold(), raw_value,
            )
        elif key == "subcategory":
            raw_value = MULTIPLE_SUBCATEGORY_NORMALIZATIONS.get(
                raw_value.casefold(), raw_value,
            )
        elif key == "vesselPart":
            raw_value = VESSEL_PART_NORMALIZATIONS.get(
                raw_value.casefold(), raw_value,
            )
        if key == "type" and get_node_datatype(node_id) == "concept-list" and re.search(r"[/,]", raw_value):
            # Prefer independent Types when every slash/comma-separated part resolves.
            # Otherwise preserve the complete source label for lookup.
            parts = [
                clean_cell(part).replace("?", "").strip()
                for part in re.split(r"\s*(?:/|,)\s*", raw_value)
                if clean_cell(part).replace("?", "").strip()
            ]
            part_ids = [
                resolve_dictionary_value(
                    dictionary,
                    apply_dictionary_alias(key, part, config.get("dictionary_aliases", {})),
                )
                for part in parts
            ]
            if parts and all(part_ids):
                return list(dict.fromkeys(part_ids))

            full_value = raw_value.replace("?", "").strip()
            full_id = resolve_dictionary_value(
                dictionary,
                apply_dictionary_alias(key, full_value, config.get("dictionary_aliases", {})),
            )
            if full_id:
                return [full_id]

            totals["unknown_concepts"] += 1
            self._record_unknown(key, full_value, source)
            return []
        if get_node_datatype(node_id) == "concept-list" and key != "type":
            separator = {
                "vesselPart": r"\s*(?:\+|/|\?\s+)\s*",
                "morphology": r"\s*(?:\+|/|,)\s*",
                "vesselForm": r"\s*(?:\+|/|,)\s*",
                "subcategory": r"\s*(?:\+|/|,)\s*",
                "provenance": r"\s*(?:\+|/|,)\s*",
            }.get(key, r"\s*(?:\+|/)\s*")
            values = [value for value in re.split(separator, raw_value) if value]
        else:

            values = [raw_value]
        if key == "provenance" and get_node_datatype(node_id) == "concept-list":
            exact_id = resolve_dictionary_value(
                dictionary,
                apply_dictionary_alias(key, raw_value, config.get("dictionary_aliases", {})),
            )
            if exact_id:
                return [exact_id]


        resolved = []
        for value in values:
            value = self._non_uncertainty_value(value)
            canonical = apply_dictionary_alias(key, value, config.get("dictionary_aliases", {}))
            value_id = resolve_dictionary_value(dictionary, canonical)
            if not value_id:
                totals["unknown_concepts"] += 1
                self._record_unknown(key, value, source)
                continue
            resolved.append(value_id)
        if get_node_datatype(node_id) == "concept-list":
            return resolved
        return resolved[0] if resolved else ""

    def _special_find_relations(self, raw_value, source, totals):
        # The fresh workbooks currently have this column empty.  Preserve a
        # warning for non-empty values until their special-find identifiers are
        # matched by a dedicated importer rather than creating an unsafe guess.
        raw_value = clean_cell(raw_value)
        if raw_value:
            totals["unknown_special_finds"] += 1
            self.stderr.write(f"  {source}: Special Find ID {raw_value!r} not linked")
        return []

    def _save_record(self, resource, record_type, record, collection_id):
        config = POTTERY_RECORD_TYPES[record_type]
        fields = config["fields"]
        extra = EXTRA_NODES[record_type]
        self._save_group(resource, {config["related_collection_node_id"]: [self._relation(collection_id)]})
        if record["formNo"]:
            self._save_group(resource, {fields["formNo"]: record["formNo"]})
        if record["pNo"]:
            self._save_group(resource, {fields["pNo"]: record["pNo"]})
        if record["count"] is not None:
            self._save_group(resource, {fields["count"]: record["count"]})
        if record["mgNo"]:
            self._save_group(resource, {extra["mgNo"]: record["mgNo"]})
        if record["author"]:
            self._save_group(resource, {extra["author"]: self._format_value(extra["author"], record["author"])})
        if record["comment"]:
            self._save_group(resource, {fields["comment"]: self._format_value(fields["comment"], record["comment"])})

        for key in (
            "vesselPart", "morphology", "vesselForm", "subcategory",
            "surfaceTreatment", "stateOfPreservation",
        ):
            node_id = extra.get(key, fields.get(key))
            value = record[key]
            if value:
                self._save_group(resource, {node_id: self._format_concept(node_id, value)})

        type_data = {}
        if record["type"]:
            type_data[fields["type"]] = self._format_value(fields["type"], record["type"])
            # The copied card requires an explicit uncertainty choice.
            type_data[fields["typeUncertain"]] = (
                record["typeUncertain"] if record["typeUncertain"] is not None else False
            )
        elif record["typeUncertain"] is not None:
            type_data[fields["typeUncertain"]] = record["typeUncertain"]
        if type_data:
            self._save_group(resource, type_data)

        chronology_data = {}
        if record["chronology"]:
            chronology_data[fields["chronology"]] = record["chronology"]
            # A blank marker in historic Excel means certain, not unknown.
            chronology_data[fields["chronologyUncertain"]] = (
                record["chronologyUncertain"]
                if record["chronologyUncertain"] is not None
                else False
            )
        elif record["chronologyUncertain"] is not None:
            chronology_data[fields["chronologyUncertain"]] = record["chronologyUncertain"]
        if record["chronologyBoundaries"]:
            chronology_data.update(self._chronology_boundary_data(
                fields["chronology"], record["chronologyBoundaries"],
            ))
        if chronology_data:
            self._save_group(resource, chronology_data)

        provenance_data = {}
        if record["provenance"]:
            provenance_data[fields["provenance"]] = self._format_concept(fields["provenance"], record["provenance"])
            provenance_data[fields["provenanceUncertain"]] = (
                record["provenanceUncertain"]
                if record["provenanceUncertain"] is not None
                else False
            )
        elif record["provenanceUncertain"] is not None:
            provenance_data[fields["provenanceUncertain"]] = record["provenanceUncertain"]
        if provenance_data:
            self._save_group(resource, provenance_data)

        for key in ("drawn", "photo"):
            if record[key] is not None:
                self._save_group(resource, {fields[key]: record[key]})

    @staticmethod
    def _format_concept(node_id, value):
        if isinstance(value, list):
            return value
        return format_concept_tile_value(node_id, value)
    @classmethod
    def _format_value(cls, node_id, value):
        datatype = get_node_datatype(node_id)
        if datatype in {"concept", "concept-list"}:
            return cls._format_concept(node_id, value)
        if datatype == "string":
            return localized_string(value)
        return value

    @staticmethod
    def _chronology_boundary_data(chronology_node_id, boundaries):
        """Return four exact EDTF chronology boundaries for one period card."""
        from arches.app.models.models import Node

        period_node = Node.objects.get(nodeid=chronology_node_id)
        nodes_by_name = {
            node.name.casefold(): str(node.nodeid)
            for node in Node.objects.filter(nodegroup_id=period_node.nodegroup_id)
        }
        required_names = (
            "earliest date", "latest start date", "earliest end date", "latest date",
        )
        missing_names = [name for name in required_names if name not in nodes_by_name]
        if missing_names:
            raise CommandError(
                "Chronology card is missing date fields: "
                + ", ".join(missing_names)
            )

        start_year, end_year = boundaries
        values = (start_year, start_year, end_year, end_year)
        return {
            nodes_by_name[name]: str(ExtendedDateFormat(str(year)).edtf)
            for name, year in zip(required_names, values)
        }

    @staticmethod
    def _save_group(resource, data):
        """Merge values for a shared card into one tile."""
        node_id = next(iter(data))
        from arches.app.models.models import Node
        node = Node.objects.get(nodeid=node_id)
        nodegroup_id = str(node.nodegroup_id)
        tile = Tile.objects.filter(
            resourceinstance_id=resource.resourceinstanceid,
            nodegroup_id=nodegroup_id,
        ).first()
        if tile is None:
            tile = Tile.get_blank_tile_from_nodegroup_id(
                nodegroup_id,
                resourceid=str(resource.resourceinstanceid),
            )
        ensure_tile_parent(resource, tile, node)
        tile.data = dict(tile.data or {})
        tile.data.update(data)
        tile.save()

    @staticmethod
    def _relation(resource_id):
        return {
            "resourceId": str(resource_id),
            "ontologyProperty": "",
            "inverseOntologyProperty": "",
            "resourceXresourceId": str(uuid4()),
        }

    def _collection_ids(self, context_label):
        key = normalize_source_context(context_label).casefold()
        if key in self.collection_ids_by_context:
            return self.collection_ids_by_context[key]
        try:
            context = self.context_importer._context_for_row({"Context": context_label})
        except (SkipRow, CommandError) as error:
            result = ([], str(error))
        else:
            resource_ids = []
            for resource_id, data in TileModel.objects.filter(
                resourceinstance__graph_id=POTTERY_GRAPH_ID,
                data__has_key=CONTEXT_NODE_ID,
            ).values_list("resourceinstance_id", "data"):
                relations = (data or {}).get(CONTEXT_NODE_ID) or []
                if any(isinstance(item, dict) and str(item.get("resourceId")) == str(context.resourceinstanceid) for item in relations):
                    resource_ids.append(resource_id)
            result = (resource_ids, "")
        self.collection_ids_by_context[key] = result
        return result

    def _create_missing_collection_if_enabled(self, context_label, collection_ids):
        """Create one minimal Pottery Collection only when the opt-in flag is set."""
        if collection_ids or not getattr(self, "create_missing_collections", False):
            return collection_ids, False, ""

        key = normalize_source_context(context_label).casefold()
        cache = getattr(self, "created_collection_ids", {})
        if key in cache:
            return cache[key], False, ""

        try:
            context = self.context_importer._context_for_row({"Context": context_label})
        except (SkipRow, CommandError) as error:
            return [], False, str(error)

        if self.apply:
            collection = Resource.objects.create(
                graph_id=POTTERY_GRAPH_ID,
                legacyid=f"{self.dataset}:collection:{key}",
            )
            self._save_group(
                collection,
                {CONTEXT_NODE_ID: [self._relation(str(context.resourceinstanceid))]},
            )
            collection.name = f"Pottery Collection {context_label}"
            collection.save(update_fields=["name"])
            ids = [str(collection.resourceinstanceid)]
        else:
            ids = [f"dry-run-collection:{key}"]

        cache[key] = ids
        self.created_collection_ids = cache
        self.collection_ids_by_context[key] = (ids, "")
        return ids, True, ""

    def _record_missing_collection(self, context_label, record_type, source, reason):
        label = normalize_source_context(context_label)
        detail = self.missing_collections.setdefault(
            label,
            {"record_types": set(), "sources": set(), "reasons": set()},
        )
        detail["record_types"].add(record_type)
        detail["sources"].add(source)
        detail["reasons"].add(reason)

    def _write_missing_collections(self):
        if not self.missing_collections:
            self.stdout.write("Missing Pottery Collections: none")
            return
        self.stdout.write("Missing Pottery Collections (unique):")
        for context_label in sorted(self.missing_collections, key=str.casefold):
            detail = self.missing_collections[context_label]
            types = ", ".join(sorted(detail["record_types"]))
            self.stdout.write(f"  {context_label}  [record types: {types}; rows: {len(detail['sources'])}]")

    @staticmethod
    def _missing_collection_report_path(options, files):
        explicit_path = clean_cell(options.get("missing_collections_report"))
        if explicit_path:
            return Path(explicit_path)
        if options.get("directory"):
            return Path(options["directory"]) / "missing_pottery_record_collections.xlsx"
        return files[0].parent / "missing_pottery_record_collections.xlsx"

    @staticmethod
    def _context_issues_report_path(options, files):
        explicit_path = clean_cell(options.get("context_issues_report") or options.get("missing_collections_report"))
        if explicit_path:
            return Path(explicit_path)
        if options.get("directory"):
            return Path(options["directory"]) / "pottery_record_context_issues.xlsx"
        return files[0].parent / "pottery_record_context_issues.xlsx"

    def _write_missing_collection_report(self):
        path = self.missing_collections_report_path
        if not path.parent.is_dir():
            raise CommandError(f"Report directory not found: {path.parent}")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Context Issues"
        sheet.append(["Context", "Issue", "Record types", "Occurrences", "Source rows"])
        for context_label in sorted(self.missing_collections, key=str.casefold):
            detail = self.missing_collections[context_label]
            sheet.append([
                context_label, "; ".join(sorted(detail["reasons"])),
                ", ".join(sorted(detail["record_types"])), len(detail["sources"]),
                "\n".join(sorted(detail["sources"])),
            ])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column, width in {"A": 28, "B": 52, "C": 24, "D": 14, "E": 96}.items():
            sheet.column_dimensions[column].width = width
        workbook.save(path)

        self.stdout.write(f"Pottery Context issues report: {path}")
    @staticmethod
    def _unknown_concepts_report_path(options, files):
        explicit_path = clean_cell(options.get("unknown_concepts_report"))
        if explicit_path:
            return Path(explicit_path)
        if options.get("directory"):
            return Path(options["directory"]) / "unknown_pottery_record_concepts.xlsx"
        return files[0].parent / "unknown_pottery_record_concepts.xlsx"

    def _record_unknown(self, field, value, source):
        value = clean_cell(value)
        if not value:
            return
        unknown_values = getattr(self, "unknown_concept_values", None)
        if unknown_values is None:
            unknown_values = self.unknown_concept_values = {}
        detail = unknown_values.setdefault((field, value), {"sources": set()})
        detail["sources"].add(source)
        if not getattr(self, "audit_unknown_concepts", False):
            self.stderr.write(f"  {source}: ignored unknown {field} value {value!r}")

    def _write_unknown_concepts(self):
        if not self.unknown_concept_values:
            self.stdout.write("Ignored unknown concepts: none")
            return
        self.stdout.write("Ignored unknown concepts (unique):")
        for (field, value), detail in sorted(self.unknown_concept_values.items(), key=lambda item: (item[0][0], item[0][1].casefold())):
            self.stdout.write(f"  {field}: {value!r}  [rows: {len(detail['sources'])}]")

    def _write_unknown_concepts_report(self):
        path = self.unknown_concepts_report_path
        if not path.parent.is_dir():
            raise CommandError(f"Report directory not found: {path.parent}")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Ignored unknown values"
        sheet.append(["Field", "Value", "Occurrences", "Source rows"])
        for (field, value), detail in sorted(self.unknown_concept_values.items(), key=lambda item: (item[0][0], item[0][1].casefold())):
            sources = sorted(detail["sources"])
            sheet.append([field, value, len(sources), "\n".join(sources)])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column, width in {"A": 22, "B": 42, "C": 14, "D": 90}.items():
            sheet.column_dimensions[column].width = width
        workbook.save(path)
        self.stdout.write(f"Unknown concepts report: {path}")

    @staticmethod
    def _type_values_report_path(options, files):
        explicit_path = clean_cell(options.get("type_values_report"))
        if explicit_path:
            return Path(explicit_path)
        if options.get("directory"):
            return Path(options["directory"]) / "plain_storage_vessel_type_values.xlsx"
        return files[0].parent / "plain_storage_vessel_type_values.xlsx"

    def _record_type_value(self, record_type, value, source):
        value = clean_cell(value)
        detail = self.type_values.setdefault((record_type, value), {"sources": set()})
        detail["sources"].add(source)

    def _write_type_values_report(self):
        path = self.type_values_report_path
        if not path.parent.is_dir():
            raise CommandError(f"Report directory not found: {path.parent}")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "PW and SV Type values"
        sheet.append(["Pottery category", "Type (as in Excel)", "Occurrences", "Source rows"])
        for (record_type, value), detail in sorted(self.type_values.items(), key=lambda item: (item[0][0], item[0][1].casefold())):
            sources = sorted(detail["sources"])
            sheet.append([POTTERY_RECORD_TYPES[record_type]["label"], value, len(sources), "\n".join(sources)])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column, width in {"A": 20, "B": 52, "C": 14, "D": 96}.items():
            sheet.column_dimensions[column].width = width
        workbook.save(path)
        self.stdout.write(f"PW/SV Type values report: {path}")

    @staticmethod
    def _has_uncertainty_marker(value):
        return "?" in clean_cell(value)

    @staticmethod
    def _non_uncertainty_value(value):
        """Strip question marks from dictionary fields without uncertainty nodes."""
        return clean_cell(value).replace("?", "").strip()

    @classmethod
    def _strip_uncertainty_marker(cls, value):
        raw_value = clean_cell(value)
        if cls._has_uncertainty_marker(raw_value):
            return cls._non_uncertainty_value(raw_value), True
        return raw_value, False
    @staticmethod
    def _number(value):

        raw_value = clean_cell(value).replace(",", ".")
        if not raw_value:
            return None
        try:
            number = float(raw_value)
        except ValueError:
            return None
        return int(number) if number.is_integer() else number

    @staticmethod
    def _value(row, field):
        for header in FIELD_HEADERS[field]:
            value = clean_cell(row.get(header))
            # Pink Excel markers occasionally survived as literal cell values.
            if value.casefold() == "ffffb6c1":
                continue
            if value:
                return value
        return ""

    def _has_record_data(self, row):
        return any(self._value(row, field) for field in FIELD_HEADERS if field not in {"formNo", "context"})

    def _legacyid(self, record_type, context_label, form_no, p_no, row_number):
        context_key = normalize_source_context(context_label)
        return f"{self.dataset}:{record_type}:{context_key}:{clean_cell(form_no)}:{clean_cell(p_no)}:{row_number}"
