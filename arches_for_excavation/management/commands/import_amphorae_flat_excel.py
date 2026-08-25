"""Import one Amphorae resource per row from a flat Excel workbook."""

from collections import Counter
from pathlib import Path
import re
from uuid import uuid4

import requests
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from arches.app.models.models import Concept, Node, Resource, TileModel, Value
from arches.app.models.tile import Tile
from arches_for_excavation.management.commands.import_collection_excel import Command as CollectionImportCommand
from arches_for_excavation.utils.pottery.common import clean_cell, localized_string, to_boolean
from arches_for_excavation.utils.pottery.constants import POTTERY_RECORD_TYPES

POTTERY_COLLECTION_GRAPH_ID = "32a4c0b9-ab8c-47a0-a42f-99cd3ad392fe"
COLLECTION_FORM_ID_NODE_ID = "25e31613-69ac-45ce-a6db-a15239de70a4"

# Keep this older flat-file command aligned with the central five-model map.
AMPHORAE_CONFIG = POTTERY_RECORD_TYPES["amphorae"]
POTTERY_RECORD_GRAPH_ID = AMPHORAE_CONFIG["graph_id"]
POTTERY_RECORD_RELATED_COLLECTION_NODE_ID = AMPHORAE_CONFIG["related_collection_node_id"]
POTTERY_RECORD_FORM_NO_NODE_ID = AMPHORAE_CONFIG["fields"]["formNo"]
POTTERY_RECORD_P_NO_NODE_ID = AMPHORAE_CONFIG["fields"]["pNo"]
POTTERY_RECORD_COUNT_NODE_ID = AMPHORAE_CONFIG["fields"]["count"]
POTTERY_RECORD_VESSEL_PART_NODE_ID = AMPHORAE_CONFIG["fields"]["vesselPart"]
POTTERY_RECORD_TYPE_NODE_ID = AMPHORAE_CONFIG["fields"]["type"]
POTTERY_RECORD_TYPE_UNCERTAIN_NODE_ID = AMPHORAE_CONFIG["fields"]["typeUncertain"]
POTTERY_RECORD_MORPHOLOGY_NODE_ID = AMPHORAE_CONFIG["fields"]["morphology"]
POTTERY_RECORD_CHRONOLOGY_NODEGROUP_ID = AMPHORAE_CONFIG["fields"]["chronology"]
POTTERY_RECORD_CHRONOLOGY_NODE_ID = AMPHORAE_CONFIG["fields"]["chronology"]
POTTERY_RECORD_CHRONOLOGY_UNCERTAIN_NODE_ID = AMPHORAE_CONFIG["fields"]["chronologyUncertain"]
POTTERY_RECORD_PROVENANCE_NODE_ID = AMPHORAE_CONFIG["fields"]["provenance"]
POTTERY_RECORD_PROVENANCE_UNCERTAIN_NODE_ID = AMPHORAE_CONFIG["fields"]["provenanceUncertain"]
POTTERY_RECORD_DRAWN_NODE_ID = AMPHORAE_CONFIG["fields"]["drawn"]
POTTERY_RECORD_PHOTO_NODE_ID = AMPHORAE_CONFIG["fields"]["photo"]
POTTERY_RECORD_COMMENT_NODE_ID = AMPHORAE_CONFIG["fields"]["comment"]

PAC_API_URL = "https://pac.cenagis.edu.pl/wiki/api.php"
LABEL_VALUE_TYPES = ("prefLabel", "altLabel", "hiddenLabel")

REQUIRED_HEADERS = {
    "Form ID",
    "P number",
    "Quantity",
    "Vessel Part",
    "Type",
    "Type Uncertainty",
    "Morphology",
    "Period",
    "Uncertain",
    "Provenance",
    "Provenance Uncertainity",
    "Drawing",
    "Photo",
    "Comment",
}

CONCEPT_FIELDS = {
    "Vessel Part": POTTERY_RECORD_VESSEL_PART_NODE_ID,
    "Type": POTTERY_RECORD_TYPE_NODE_ID,
    "Morphology": POTTERY_RECORD_MORPHOLOGY_NODE_ID,
    "Provenance": POTTERY_RECORD_PROVENANCE_NODE_ID,
}


class Command(BaseCommand):
    help = "Import a flat Amphorae workbook; each source row becomes one resource."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True)
        parser.add_argument("--sheet", help="Sheet name; defaults to the first sheet.")
        parser.add_argument(
            "--dataset",
            help="Stable source name used for idempotency; defaults to the workbook stem.",
        )
        parser.add_argument("--apply", action="store_true")
        parser.add_argument(
            "--sync-pac",
            action="store_true",
            help=(
                "Resolve missing concept labels against the PAC API and create their "
                "local Concept/Value rows. This sends unique dictionary labels to PAC."
            ),
        )

    def handle(self, *args, **options):
        path = Path(options["file"])
        if not path.is_file():
            raise CommandError(f"Workbook not found: {path}")

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[options["sheet"]] if options.get("sheet") else workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [clean_cell(value) for value in next(rows)]
        except StopIteration as error:
            raise CommandError("Workbook is empty.") from error
        missing_headers = REQUIRED_HEADERS - set(headers)
        if missing_headers:
            raise CommandError(f"Missing columns: {', '.join(sorted(missing_headers))}")

        dataset = clean_cell(options.get("dataset")) or path.stem
        apply = options["apply"]
        sync_pac = options["sync_pac"]
        source_rows = [
            (row_number, dict(zip(headers, values)))
            for row_number, values in enumerate(rows, start=2)
            if any(clean_cell(value) for value in values)
        ]

        self.stdout.write(f"Flat Amphorae import [{'APPLY' if apply else 'DRY-RUN'}]")
        self.stdout.write(
            f"Workbook: {path.name}; sheet: {sheet.title}; rows: {len(source_rows)}; "
            f"dataset: {dataset}"
        )

        prepared, preparation_totals = self._prepare_rows(
            source_rows,
            dataset,
            apply=apply,
            sync_pac=sync_pac,
        )
        unresolved = preparation_totals["unresolved_concepts"]
        if unresolved:
            self.stderr.write(
                self.style.WARNING(
                    f"Unresolved concept cells: {unresolved}. They will be omitted; "
                    "the source Comment is preserved."
                )
            )

        totals = Counter(preparation_totals)
        for item in prepared:
            outcome = self._process_row(item, apply=apply)
            totals[outcome] += 1

        self.stdout.write("Summary:")
        for key in (
            "created",
            "would_create",
            "unchanged",
            "missing_collections",
            "duplicate_collections",
            "invalid_quantity",
            "unresolved_concepts",
        ):
            self.stdout.write(f"  {key}: {totals[key]}")
        if not apply:
            self.stdout.write("Dry-run only. Add --apply to create resources.")

    def _prepare_rows(self, source_rows, dataset, apply, sync_pac):
        totals = Counter()
        concept_cache = {}
        prepared = []

        for row_number, row in source_rows:
            form_id = clean_cell(row.get("Form ID"))
            item = {
                "row_number": row_number,
                "row": row,
                "form_id": form_id,
                "legacyid": f"amphorae-flat:{dataset}:{sheet_safe(row_number)}",
                "concepts": {},
            }

            for field in CONCEPT_FIELDS:
                label = clean_cell(row.get(field))
                if not label:
                    continue
                if field == "Vessel Part" and "+" in label:
                    vessel_part_ids = []
                    for vessel_part in (
                        clean_cell(value) for value in re.split(r"\s*\+\s*", label)
                    ):
                        if not vessel_part:
                            continue
                        value_id = self._concept_value_id(
                            vessel_part,
                            concept_cache,
                            apply=apply,
                            sync_pac=sync_pac,
                        )
                        if value_id:
                            # Do not deduplicate: "body + body" means two Body tiles.
                            vessel_part_ids.append(value_id)
                        else:
                            totals["unresolved_concepts"] += 1
                            self.stderr.write(
                                self.style.WARNING(
                                    f"  row {row_number}: unresolved {field} "
                                    f"component {vessel_part!r}"
                                )
                            )
                    if vessel_part_ids:
                        item["concepts"][field] = vessel_part_ids
                    continue

                value_id = self._concept_value_id(
                    label,
                    concept_cache,
                    apply=apply,
                    sync_pac=sync_pac,
                )
                if value_id:
                    item["concepts"][field] = value_id
                else:
                    totals["unresolved_concepts"] += 1
                    self.stderr.write(
                        self.style.WARNING(
                            f"  row {row_number}: unresolved {field} {label!r}"
                        )
                    )

            raw_quantity = clean_cell(row.get("Quantity"))
            if raw_quantity and self._number(raw_quantity) is None:
                totals["invalid_quantity"] += 1
                self.stderr.write(
                    self.style.WARNING(
                        f"  row {row_number}: invalid Quantity {raw_quantity!r}; omitted"
                    )
                )

            unknown_periods = []
            period_ids = CollectionImportCommand._period_values(
                row.get("Period"),
                unknown_periods,
            )
            if unknown_periods:
                label = clean_cell(row.get("Period"))
                value_id = ""
                # Preserve locally resolved parts of a range. Only try the full
                # label as a fallback when none of it was understood.
                if not period_ids:
                    cache_key = f"chronology:{label.casefold()}"
                    if cache_key not in concept_cache:
                        if sync_pac:
                            value_id = self._sync_pac_value(label, apply=apply)
                        concept_cache[cache_key] = value_id
                    value_id = concept_cache[cache_key]
                if value_id:
                    period_ids.append(value_id)
                else:
                    totals["unresolved_concepts"] += len(unknown_periods)
                    for unknown_period in unknown_periods:
                        self.stderr.write(
                            self.style.WARNING(
                                f"  row {row_number}: unresolved Period tail "
                                f"{unknown_period!r}; keeping recognized period(s)"
                            )
                        )
            item["period_ids"] = list(dict.fromkeys(period_ids))
            prepared.append(item)

        return prepared, totals

    def _process_row(self, item, apply):
        legacyid = item["legacyid"]
        if Resource.objects.filter(graph_id=POTTERY_RECORD_GRAPH_ID, legacyid=legacyid).exists():
            self.stdout.write(f"  row {item['row_number']}: unchanged")
            return "unchanged"

        collection_ids = list(
            TileModel.objects.filter(
                resourceinstance__graph_id=POTTERY_COLLECTION_GRAPH_ID,
                data__contains={COLLECTION_FORM_ID_NODE_ID: item["form_id"]},
            )
            .values_list("resourceinstance_id", flat=True)
            .distinct()
        )
        if not collection_ids:
            self.stdout.write(
                f"  row {item['row_number']}: skipped "
                f"{item['form_id']} (collection not found)"
            )
            return "missing_collections"
        if len(collection_ids) != 1:
            self.stderr.write(
                f"  row {item['row_number']}: {item['form_id']} has "
                f"{len(collection_ids)} collections"
            )
            return "duplicate_collections"

        if not apply:
            self.stdout.write(
                f"  row {item['row_number']}: would create Amphorae Pottery Record for {item['form_id']}"
            )
            return "would_create"

        with transaction.atomic():
            resource = self._create_resource(item, str(collection_ids[0]))
        self.stdout.write(f"  row {item['row_number']}: created {resource.resourceinstanceid}")
        return "created"

    def _create_resource(self, item, collection_resource_id):
        row = item["row"]
        resource = Resource.objects.create(
            graph_id=POTTERY_RECORD_GRAPH_ID,
            legacyid=item["legacyid"],
            name=f"Amphorae {item['form_id']} row {item['row_number']}",
        )

        self._save_node(
            resource,
            POTTERY_RECORD_RELATED_COLLECTION_NODE_ID,
            [{
                "resourceId": collection_resource_id,
                "ontologyProperty": "",
                "inverseOntologyProperty": "",
                "resourceXresourceId": str(uuid4()),
            }],
        )
        self._save_node(resource, POTTERY_RECORD_FORM_NO_NODE_ID, item["form_id"])

        p_number = clean_cell(row.get("P number"))
        if p_number:
            self._save_node(resource, POTTERY_RECORD_P_NO_NODE_ID, p_number)

        quantity = self._number(row.get("Quantity"))
        if quantity is not None:
            self._save_node(resource, POTTERY_RECORD_COUNT_NODE_ID, quantity)

        vessel_part_ids = item["concepts"].get("Vessel Part") or []
        if not isinstance(vessel_part_ids, list):
            vessel_part_ids = [vessel_part_ids]
        if vessel_part_ids:
            # Vessel Part is a concept-list: one tile can retain repeated values,
            # so "body + body" is stored as [body_id, body_id].
            self._save_node(
                resource,
                POTTERY_RECORD_VESSEL_PART_NODE_ID,
                vessel_part_ids,
            )

        type_data = {}
        if item["concepts"].get("Type"):
            type_data[POTTERY_RECORD_TYPE_NODE_ID] = [item["concepts"]["Type"]]
        type_data[POTTERY_RECORD_TYPE_UNCERTAIN_NODE_ID] = to_boolean(
            row.get("Type Uncertainty")
        )
        self._save_group(resource, POTTERY_RECORD_TYPE_NODE_ID, type_data)

        if item["concepts"].get("Morphology"):
            self._save_node(
                resource,
                POTTERY_RECORD_MORPHOLOGY_NODE_ID,
                [item["concepts"]["Morphology"]],
            )

        chronology_data = {
            POTTERY_RECORD_CHRONOLOGY_UNCERTAIN_NODE_ID: to_boolean(row.get("Uncertain")),
        }
        if item["period_ids"]:
            chronology_data[POTTERY_RECORD_CHRONOLOGY_NODE_ID] = item["period_ids"]
        self._save_group(resource, POTTERY_RECORD_CHRONOLOGY_NODEGROUP_ID, chronology_data)

        provenance_data = {
            POTTERY_RECORD_PROVENANCE_UNCERTAIN_NODE_ID: to_boolean(
                row.get("Provenance Uncertainity")
            ),
        }
        if item["concepts"].get("Provenance"):
            provenance_data[POTTERY_RECORD_PROVENANCE_NODE_ID] = [item["concepts"]["Provenance"]]
        self._save_group(resource, POTTERY_RECORD_PROVENANCE_NODE_ID, provenance_data)

        self._save_node(resource, POTTERY_RECORD_DRAWN_NODE_ID, to_boolean(row.get("Drawing")))
        self._save_node(resource, POTTERY_RECORD_PHOTO_NODE_ID, to_boolean(row.get("Photo")))

        comment = clean_cell(row.get("Comment"))
        if comment:
            self._save_node(resource, POTTERY_RECORD_COMMENT_NODE_ID, localized_string(comment))

        resource.name = f"Amphorae {item['form_id']} row {item['row_number']}"
        resource.save(update_fields=["name"])
        return resource

    @staticmethod
    def _save_node(resource, node_id, value):
        Command._save_group(resource, node_id, {node_id: value})

    @staticmethod
    def _save_group(resource, node_or_nodegroup_id, data):
        """Save all values from one card into one tile on the copied model."""
        node = Node.objects.get(nodeid=node_or_nodegroup_id)
        nodegroup_id = str(node.nodegroup_id)
        tile = Tile.objects.filter(
            resourceinstance_id=resource.resourceinstanceid,
            nodegroup_id=nodegroup_id,
        ).first()
        if tile is None:
            tile = Tile.get_blank_tile_from_nodegroup_id(
                nodegroup_id,
                resourceid=str(resource.resourceinstanceid),
            )
        tile.data = dict(tile.data or {})
        tile.data.update(data)
        tile.save()
        return tile

    @staticmethod
    def _number(value):
        raw = clean_cell(value)
        if not raw:
            return None
        try:
            number = float(raw)
        except ValueError:
            return None
        return int(number) if number.is_integer() else number

    @staticmethod
    def _local_concept_value_id(label):
        values = Value.objects.filter(
            value__iexact=clean_cell(label),
            valuetype_id__in=LABEL_VALUE_TYPES,
        ).select_related("concept")
        concept_ids = {str(value.concept_id) for value in values}
        if len(concept_ids) != 1:
            return ""
        preferred = Value.objects.filter(
            concept_id=next(iter(concept_ids)),
            valuetype_id="prefLabel",
        ).order_by("language_id")
        value = preferred.filter(language_id="en").first() or preferred.first()
        return str(value.valueid) if value else ""

    def _concept_value_id(self, label, cache, apply, sync_pac):
        cache_key = f"concept:{clean_cell(label).casefold()}"
        if cache_key not in cache:
            value_id = self._local_concept_value_id(label)
            if not value_id and sync_pac:
                value_id = self._sync_pac_value(label, apply=apply)
            cache[cache_key] = value_id
        return cache[cache_key]

    def _sync_pac_value(self, label, apply):
        try:
            response = requests.get(
                PAC_API_URL,
                params={
                    "action": "wbsearchentities",
                    "search": clean_cell(label),
                    "language": "en",
                    "format": "json",
                    "limit": 20,
                },
                timeout=20,
            )
            response.raise_for_status()
        except requests.RequestException as error:
            self.stderr.write(f"PAC lookup failed for {label!r}: {error}")
            return ""

        exact = [
            result
            for result in response.json().get("search", [])
            if clean_cell(result.get("label")).casefold() == clean_cell(label).casefold()
            or clean_cell((result.get("match") or {}).get("text")).casefold()
            == clean_cell(label).casefold()
        ]
        qids = {result.get("id") for result in exact if result.get("id")}
        if len(qids) != 1:
            return ""
        qid = next(iter(qids))
        canonical_label = next(
            (clean_cell(result.get("label")) for result in exact if result.get("label")),
            clean_cell(label),
        )
        legacyoid = f"https://pac.cenagis.edu.pl/entity/{qid}"
        existing = Concept.objects.filter(legacyoid=legacyoid).first()
        if existing:
            preferred = Value.objects.filter(
                concept=existing,
                valuetype_id="prefLabel",
            ).order_by("language_id")
            value = preferred.filter(language_id="en").first() or preferred.first()
            return str(value.valueid) if value else ""
        if not apply:
            return f"PAC:{qid}"

        concept = Concept.objects.create(nodetype_id="Concept", legacyoid=legacyoid)
        value = Value.objects.create(
            concept=concept,
            valuetype_id="prefLabel",
            value=canonical_label,
            language_id="en",
        )
        return str(value.valueid)


def sheet_safe(row_number):
    return str(int(row_number))
