"""Import remarks from ceramika/audyt using Form ID → full Context mapping."""

from collections import Counter
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from arches_for_excavation.management.commands.import_collection_excel import Command as CollectionImportCommand
from arches_for_excavation.management.commands.import_pottery_remarks_excel import (
    CATEGORY_TYPES,
    FILENAME_CATEGORY,
    FORM_ID_HEADERS,
    REMARKS_HEADER,
    Command as RemarksCommand,
)
from arches_for_excavation.utils.pottery.common import clean_cell


class Command(BaseCommand):
    help = "Import ceramika/audyt remarks using full Context values from source category workbooks."

    def add_arguments(self, parser):
        parser.add_argument("--directory", required=True, help="ceramika directory containing audyt and source workbooks.")
        parser.add_argument("--apply", action="store_true", help="Save updates. Default is dry-run.")

    def handle(self, *args, **options):
        directory = Path(options["directory"])
        if not directory.is_dir():
            raise CommandError(f"Directory not found: {directory}")
        audit_files = sorted(path for path in (directory / "audyt").rglob("*.xlsx") if not path.name.startswith("~$"))
        if not audit_files:
            raise CommandError(f"No audit workbooks found in {directory / 'audyt'}")

        importer = RemarksCommand()
        importer.apply = options["apply"]
        importer.create_missing_collections = False
        importer.context_importer = CollectionImportCommand()
        importer.context_importer.apply = importer.apply
        importer.context_importer.create_missing_contexts = False
        importer.context_importer.context_resource_id = None
        importer.context_importer.context_map = {}
        importer.context_importer.created_contexts = {}
        importer.context_importer.trench_resource_ids = None
        importer.category_value_ids = {category: importer._category_value_id(qid) for category, qid in CATEGORY_TYPES.items()}
        importer.seen_rows = {}
        importer.collection_ids_by_context = {}

        form_contexts, map_totals = self._form_contexts(directory)
        totals = Counter(map_totals)
        self.stdout.write(f"Audit pottery remarks import [{'APPLY' if importer.apply else 'DRY-RUN'}]")
        self.stdout.write(f"Form ID → Context mappings: {len(form_contexts)}")

        for path in audit_files:
            category_match = FILENAME_CATEGORY.search(path.stem)
            if not category_match:
                totals["errors"] += 1
                self.stderr.write(f"Cannot derive category from {path.name}")
                continue
            category = category_match.group(1).upper()
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    rows = sheet.iter_rows(values_only=False)
                    try:
                        headers = [clean_cell(cell.value) for cell in next(rows)]
                    except StopIteration:
                        continue
                    form_header = next((header for header in FORM_ID_HEADERS if header in headers), None)
                    if not form_header or REMARKS_HEADER not in headers:
                        continue
                    self.stdout.write(f"Workbook: {path.name}; sheet: {sheet.title} ({category})")
                    for row_number, cells in enumerate(rows, start=2):
                        values = [cell.value for cell in cells]
                        row = dict(zip(headers, values))
                        if RemarksCommand._context_is_marked_red(cells, headers):
                            totals["skipped_red_context"] += 1
                            self.stdout.write(f"  {path.name}:{sheet.title}:{row_number}: skipped Context marked red")
                            continue
                        form_id = clean_cell(row.get(form_header))
                        remarks = clean_cell(row.get(REMARKS_HEADER))
                        if not form_id or not remarks:
                            continue
                        context = form_contexts.get((category, form_id))
                        if not context:
                            totals["missing_source_context"] += 1
                            self.stdout.write(f"  {path.name}:{sheet.title}:{row_number}: no full Context for {form_id}")
                            continue
                        totals.update(importer._process_row(path, sheet.title, row_number, form_id, category, remarks, context))
            finally:
                workbook.close()

        self.stdout.write("Summary:")
        for key in ("source_rows", "source_conflicts", "source_skipped_red_context", "rows_seen", "skipped_red_context", "updated", "unchanged", "missing_source_context", "missing_collections", "missing_fragments", "duplicate_rows", "conflicting_rows", "errors"):
            self.stdout.write(f"  {key}: {totals[key]}")
        if not importer.apply:
            self.stdout.write("Dry-run only. Add --apply to save changes.")

    @staticmethod
    def _is_marked_red(cell):
        color = cell.fill.fgColor
        return color.type == "rgb" and (color.rgb or "").upper().endswith("FFB6C1")

    @staticmethod
    def _form_contexts(directory):
        mappings = {}
        totals = Counter()
        ambiguous = set()
        for path in directory.rglob("*.xlsx"):
            if path.name.startswith("~$") or "audyt" in {part.casefold() for part in path.parts}:
                continue
            category_match = FILENAME_CATEGORY.search(path.stem)
            if not category_match:
                continue
            category = category_match.group(1).upper()
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    rows = sheet.iter_rows(values_only=False)
                    try:
                        headers = [clean_cell(cell.value) for cell in next(rows)]
                    except StopIteration:
                        continue
                    form_header = next((header for header in FORM_ID_HEADERS if header in headers), None)
                    if not form_header or "Pottery Collection" not in headers:
                        continue
                    context_index = headers.index("Pottery Collection")
                    for cells in rows:
                        values = [cell.value for cell in cells]
                        row = dict(zip(headers, values))
                        if Command._is_marked_red(cells[context_index]):
                            totals["source_skipped_red_context"] += 1
                            continue
                        form_id = clean_cell(row.get(form_header))
                        context = clean_cell(row.get("Pottery Collection"))
                        if not form_id or not context:
                            continue
                        key = (category, form_id)
                        if key in ambiguous:
                            continue
                        prior = mappings.get(key)
                        if prior and prior != context:
                            totals["source_conflicts"] += 1
                            mappings.pop(key, None)
                            ambiguous.add(key)
                            continue
                        mappings[key] = context
                        totals["source_rows"] += 1
            finally:
                workbook.close()
        return mappings, totals
