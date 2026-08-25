"""Import Pottery Category Form remarks from all pottery Excel workbooks.

The ``ceramika`` directory contains two source layouts:

* a separate ``A_*_m_popr.xlsx``, ``TW_*_m_popr.xlsx`` etc. workbook for one
  pottery category;
* a ``Main_*`` workbook with one remarks column per category.

This command recognises both layouts and updates the matching category tile in
an existing Pottery Collection.  It is deliberately dry-run by default.
"""

from collections import Counter
from pathlib import Path
import re
from uuid import uuid4

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import Workbook, load_workbook

from arches.app.models.models import Resource, TileModel, Value
from arches_for_excavation.management.commands.import_collection_excel import (
    Command as CollectionImportCommand,
    SkipRow,
)
from arches_for_excavation.management.commands.import_collection_excel import (
    CONTEXT_NODEGROUP_ID,
    CONTEXT_NODE_ID,
    FORM_ID_NODEGROUP_ID,
)
from arches_for_excavation.utils.pottery.common import clean_cell, localized_string


POTTERY_GRAPH_ID = "32a4c0b9-ab8c-47a0-a42f-99cd3ad392fe"
FORM_ID_NODE_ID = "25e31613-69ac-45ce-a6db-a15239de70a4"
FRAGMENT_NODEGROUP_ID = "8f7a5ca4-9c49-405d-9a08-a8debb13a9ec"
POTTERY_TYPE_NODE_ID = "3bc235a3-2240-4e94-b8af-f4c70ee13af0"
CATEGORY_REMARKS_NODE_ID = "3c371503-9028-464a-8b85-53a43c853781"
PAC_ENTITY_URL = "https://pac.cenagis.edu.pl/entity/"

FORM_ID_HEADERS = ("Form ID", "Form_ID")
REMARKS_HEADER = "Remarks (from Pottery Category Form)"

# The source abbreviations are also the prefixes used in the individual
# category workbook names.  Main workbooks additionally contain Kitchen Ware.
CATEGORY_TYPES = {
    "TW": "Q937",
    "A": "Q969",
    "KW": "Q970",
    "PW": "Q938",
    "SV": "Q971",
    "L": "Q924",
}
MAIN_REMARKS_HEADERS = {
    f"{category}_Remarks_from_Pottery_Category_Form": category
    for category in CATEGORY_TYPES
}
FILENAME_CATEGORY = re.compile(r"(?:^|[_-])(TW|SV|PW|KW|A|L)(?:[_-]|$)", re.IGNORECASE)


class Command(BaseCommand):
    help = (
        "Import Pottery Category Form remarks from category and Main Excel "
        "workbooks in a directory."
    )

    def add_arguments(self, parser):
        source = parser.add_mutually_exclusive_group(required=True)
        source.add_argument(
            "--directory",
            help="Directory containing .xlsx pottery workbooks (searched recursively).",
        )
        source.add_argument(
            "--file",
            action="append",
            dest="files",
            help="One .xlsx workbook to import; may be supplied more than once.",
        )
        parser.add_argument(
            "--apply",
            action="store_true",
            help="Save updates. Without this flag the command is a dry-run.",
        )
        parser.add_argument(
            "--create-missing-collections",
            action="store_true",
            help=(
                "Create a minimal Pottery Collection, its Context and missing "
                "Pottery Fragment when no collection exists."
            ),

        )
        parser.add_argument(
            "--missing-collections-report",
            help="Optional .xlsx path for the dry-run list of missing Pottery Collections.",
        )

    def handle(self, *args, **options):
        files = self._source_files(options)
        if not files:
            raise CommandError("No .xlsx workbooks found.")

        self.apply = options["apply"]
        self.create_missing_collections = options["create_missing_collections"]
        self.missing_collections_report_path = self._missing_collection_report_path(options, files)
        self.context_importer = CollectionImportCommand()
        self.context_importer.apply = self.apply
        self.context_importer.create_missing_contexts = self.create_missing_collections
        self.context_importer.context_resource_id = None
        self.context_importer.context_map = {}
        self.context_importer.created_contexts = {}
        self.context_importer.trench_resource_ids = None
        self.category_value_ids = {
            category: self._category_value_id(qid)
            for category, qid in CATEGORY_TYPES.items()
        }
        self.seen_rows = {}
        self.collection_ids_by_context = {}
        self.missing_collection_details = {}
        totals = Counter()

        self.stdout.write(
            f"Pottery remarks import [{'APPLY' if self.apply else 'DRY-RUN'}]"
        )
        self.stdout.write(f"Workbooks: {len(files)}")

        for path in files:
            report = self._process_workbook(path)
            totals.update(report)

        self.stdout.write("Summary:")
        for key in (
            "workbooks_recognized",
            "workbooks_skipped",
            "rows_seen",
            "updated",
            "unchanged",
            "skipped_empty",
            "skipped_red_context",
            "missing_collections",
            "created_collections",
            "would_create_collections",
            "created_fragments",
            "would_create_fragments",
            "missing_fragments",
            "duplicate_rows",
            "conflicting_rows",
            "errors",
        ):
            self.stdout.write(f"  {key}: {totals[key]}")
        if not self.apply:
            self._write_missing_collection_list()
            self._write_missing_collection_report()
        if not self.apply:
            self.stdout.write("Dry-run only. Add --apply to save changes.")

    @staticmethod
    def _context_is_marked_red(cells, headers):
        """The pink Context fill in the source sheets means 'remove / skip'."""
        if "Context" not in headers:
            return False
        cell = cells[headers.index("Context")]
        color = cell.fill.fgColor
        return color.type == "rgb" and (color.rgb or "").upper().endswith("FFB6C1")

    def _record_missing_collection(self, context_label, category, source):
        label = clean_cell(context_label)
        if not label:
            label = "(empty Context)"
        detail = self.missing_collection_details.setdefault(
            label,
            {"categories": set(), "sources": set()},
        )
        detail["categories"].add(category)
        detail["sources"].add(source)

    def _write_missing_collection_list(self):
        if not self.missing_collection_details:
            self.stdout.write("Missing Pottery Collections: none")
            return
        self.stdout.write("Missing Pottery Collections (unique; non-empty remarks only):")
        for label in sorted(self.missing_collection_details, key=str.casefold):
            detail = self.missing_collection_details[label]
            categories = ", ".join(sorted(detail["categories"]))
            count = len(detail["sources"])
            self.stdout.write(f"  {label}  [categories: {categories}; rows: {count}]")

    @staticmethod
    def _missing_collection_report_path(options, files):
        explicit_path = clean_cell(options.get("missing_collections_report"))
        if explicit_path:
            return Path(explicit_path)
        if options.get("directory"):
            return Path(options["directory"]) / "missing_pottery_remarks_collections.xlsx"
        return files[0].parent / "missing_pottery_remarks_collections.xlsx"

    def _write_missing_collection_report(self):
        path = self.missing_collections_report_path
        if not path.parent.is_dir():
            raise CommandError(f"Report directory not found: {path.parent}")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Missing Collections"
        sheet.append(["Context", "Categories", "Source rows"])
        for label in sorted(self.missing_collection_details, key=str.casefold):
            detail = self.missing_collection_details[label]
            sheet.append([label, ", ".join(sorted(detail["categories"])), len(detail["sources"])])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column, width in {"A": 30, "B": 24, "C": 14}.items():
            sheet.column_dimensions[column].width = width
        workbook.save(path)
        self.stdout.write(f"Missing Pottery Collections report: {path}")
    @staticmethod
    def _source_files(options):
        if options.get("directory"):
            directory = Path(options["directory"])
            if not directory.is_dir():
                raise CommandError(f"Directory not found: {directory}")
            return sorted(
                path for path in directory.rglob("*.xlsx")
                if not path.name.startswith("~$")
                and "audyt" not in {part.casefold() for part in path.parts}
            )

        paths = [Path(value) for value in options.get("files", [])]
        missing = [str(path) for path in paths if not path.is_file()]
        if missing:
            raise CommandError(f"Workbook not found: {', '.join(missing)}")
        return paths

    def _process_workbook(self, path):
        workbook = load_workbook(path, read_only=True, data_only=True)
        totals = Counter()
        recognized = False

        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=False)
            try:
                headers = [clean_cell(cell.value) for cell in next(rows)]
            except StopIteration:
                continue

            form_header = next((header for header in FORM_ID_HEADERS if header in headers), None)
            if not form_header:
                continue

            main_categories = {
                header: category
                for header, category in MAIN_REMARKS_HEADERS.items()
                if header in headers
            }
            if main_categories:
                recognized = True
                self.stdout.write(f"Workbook: {path.name}; sheet: {sheet.title} (Main)")
                for row_number, cells in enumerate(rows, start=2):
                    values = [cell.value for cell in cells]
                    row = dict(zip(headers, values))
                    if self._context_is_marked_red(cells, headers):
                        totals["skipped_red_context"] += 1
                        self.stdout.write(f"  row {row_number}: skipped Context marked red")
                        continue
                    form_id = clean_cell(row.get(form_header))
                    if not form_id:
                        if any(clean_cell(value) for value in values):
                            totals["errors"] += 1
                            self.stderr.write(f"  row {row_number}: missing Form ID")
                        continue
                    for header, category in main_categories.items():
                        totals.update(
                            self._process_row(
                                path, sheet.title, row_number, form_id, category,
                                clean_cell(row.get(header)),
                                self._row_context_label(row),
                            )
                        )
                continue

            if REMARKS_HEADER not in headers:
                continue
            category = self._category_from_filename(path)
            if not category:
                totals["errors"] += 1
                self.stderr.write(
                    self.style.ERROR(
                        f"Workbook {path.name}: remarks column found, but pottery "
                        "category cannot be derived from its filename."
                    )
                )
                continue

            recognized = True
            self.stdout.write(
                f"Workbook: {path.name}; sheet: {sheet.title} ({category})"
            )
            for row_number, cells in enumerate(rows, start=2):
                values = [cell.value for cell in cells]
                row = dict(zip(headers, values))
                if self._context_is_marked_red(cells, headers):
                    totals["skipped_red_context"] += 1
                    self.stdout.write(f"  row {row_number}: skipped Context marked red")
                    continue
                form_id = clean_cell(row.get(form_header))
                remarks = clean_cell(row.get(REMARKS_HEADER))
                if not form_id and not remarks:
                    continue
                if not form_id:
                    totals["errors"] += 1
                    self.stderr.write(f"  row {row_number}: missing Form ID")
                    continue
                totals.update(
                    self._process_row(
                        path, sheet.title, row_number, form_id, category,
                        remarks, self._row_context_label(row),
                    )
                )

        totals["workbooks_recognized" if recognized else "workbooks_skipped"] += 1
        return totals

    def _process_row(self, path, sheet_name, row_number, form_id, category, remarks, context_label=""):
        totals = Counter(rows_seen=1)
        if not remarks:
            totals["skipped_empty"] += 1
            return totals

        row_key = (clean_cell(context_label).casefold() or form_id, category)
        prior = self.seen_rows.get(row_key)
        source = f"{path.name}:{sheet_name}:{row_number}"
        if prior:
            if prior["remarks"] == remarks:
                totals["duplicate_rows"] += 1
                return totals
            totals["conflicting_rows"] += 1
            self.stderr.write(
                self.style.ERROR(
                    f"  {source}: conflicting {category} remarks for {form_id}; "
                    f"first encountered at {prior['source']}."
                )
            )
            return totals
        self.seen_rows[row_key] = {"remarks": remarks, "source": source}

        resource_ids = self._collection_ids_for_context(context_label)
        if not resource_ids:
            if not self.create_missing_collections:
                totals["missing_collections"] += 1
                self._record_missing_collection(context_label, category, source)
                self.stdout.write(f"  {source}: skipped {context_label!r} (collection not found)")
                return totals
            return self._create_missing_collection(form_id, category, remarks, context_label, source, totals)
        if len(resource_ids) != 1:
            totals["errors"] += 1
            self.stderr.write(
                f"  {source}: Context {context_label!r} has {len(resource_ids)} collections"
            )
            return totals

        fragments = list(
            TileModel.objects.filter(
                resourceinstance_id=resource_ids[0],
                nodegroup_id=FRAGMENT_NODEGROUP_ID,
                data__contains={POTTERY_TYPE_NODE_ID: self.category_value_ids[category]},
            )
        )
        if not fragments:
            if not self.create_missing_collections:
                totals["missing_fragments"] += 1
                self.stdout.write(
                    f"  {source}: skipped {form_id} ({category} fragment not found)"
                )
                return totals
            return self._create_missing_fragment(resource_ids[0], form_id, category, remarks, source, totals)
        if len(fragments) != 1:
            totals["errors"] += 1
            self.stderr.write(
                f"  {source}: {form_id} has {len(fragments)} {category} fragments"
            )
            return totals

        fragment = fragments[0]
        new_value = localized_string(remarks)
        if fragment.data.get(CATEGORY_REMARKS_NODE_ID) == new_value:
            totals["unchanged"] += 1
            return totals

        if self.apply:
            with transaction.atomic():
                fragment.data[CATEGORY_REMARKS_NODE_ID] = new_value
                fragment.save()
            self.stdout.write(f"  {source}: updated {form_id} ({category})")
        else:
            self.stdout.write(f"  {source}: would update {form_id} ({category})")
        totals["updated"] += 1
        return totals

    def _collection_ids_for_context(self, context_label):
        key = clean_cell(context_label).casefold()
        if not key:
            return []
        if key in self.collection_ids_by_context:
            return self.collection_ids_by_context[key]
        original_create_flag = self.context_importer.create_missing_contexts
        self.context_importer.create_missing_contexts = False
        try:
            context = self.context_importer._context_for_row({"Context": context_label})
        except (SkipRow, CommandError):
            resource_ids = []
        else:
            resource_ids = []
            tiles = (
                TileModel.objects.filter(
                    resourceinstance__graph_id=POTTERY_GRAPH_ID,
                    data__has_key=CONTEXT_NODE_ID,
                )
                .values_list("resourceinstance_id", "data")
            )
            for resource_id, data in tiles:
                relations = (data or {}).get(CONTEXT_NODE_ID) or []
                if any(
                    isinstance(relation, dict)
                    and str(relation.get("resourceId")) == str(context.resourceinstanceid)
                    for relation in relations
                ):
                    resource_ids.append(resource_id)
        finally:
            self.context_importer.create_missing_contexts = original_create_flag
        self.collection_ids_by_context[key] = resource_ids
        return resource_ids

    @staticmethod
    def _row_context_label(row):
        """Return the source context label used to locate a collection.

        ``Main`` and ``*_m_*`` sheets call it ``Context``; individual pottery
        sheets call the same identifier ``Pottery Collection``.
        """
        return clean_cell(row.get("Pottery Collection") or row.get("Context"))

    def _create_missing_collection(self, form_id, category, remarks, context_label, source, totals):
        if not context_label:
            totals["errors"] += 1
            self.stderr.write(f"  {source}: cannot create {form_id} (missing Context)")
            return totals
        try:
            if self.apply:
                with transaction.atomic():
                    context = self.context_importer._context_for_row({"Context": context_label})
                    resource = Resource.objects.create(graph_id=POTTERY_GRAPH_ID)
                    context_data = {
                        CONTEXT_NODE_ID: [{
                            "resourceId": str(context.resourceinstanceid),
                            "ontologyProperty": "",
                            "inverseOntologyProperty": "",
                            "resourceXresourceId": str(uuid4()),
                        }],
                    }
                    CollectionImportCommand._save_tile(resource, CONTEXT_NODEGROUP_ID, context_data)
                    CollectionImportCommand._save_tile(resource, FORM_ID_NODEGROUP_ID, {FORM_ID_NODE_ID: form_id})
                    CollectionImportCommand._save_tile(resource, FRAGMENT_NODEGROUP_ID, {
                        POTTERY_TYPE_NODE_ID: self.category_value_ids[category],
                        CATEGORY_REMARKS_NODE_ID: localized_string(remarks),
                    })
            else:
                self.context_importer._context_for_row({"Context": context_label})
        except (SkipRow, CommandError) as error:
            totals["errors"] += 1
            self.stderr.write(f"  {source}: cannot create {form_id} ({error})")
            return totals
        if self.apply:
            totals["created_collections"] += 1
            self.stdout.write(f"  {source}: created {form_id} ({category})")
        else:
            totals["would_create_collections"] += 1
            self.stdout.write(f"  {source}: would create {form_id} ({category})")
        return totals

    def _create_missing_fragment(self, resource_id, form_id, category, remarks, source, totals):
        if self.apply:
            with transaction.atomic():
                resource = Resource.objects.get(resourceinstanceid=resource_id)
                CollectionImportCommand._save_tile(resource, FRAGMENT_NODEGROUP_ID, {POTTERY_TYPE_NODE_ID: self.category_value_ids[category], CATEGORY_REMARKS_NODE_ID: localized_string(remarks)})
            totals["created_fragments"] += 1
            self.stdout.write(f"  {source}: created {category} fragment for {form_id}")
        else:
            totals["would_create_fragments"] += 1
            self.stdout.write(f"  {source}: would create {category} fragment for {form_id}")
        return totals

    @staticmethod
    def _category_from_filename(path):
        match = FILENAME_CATEGORY.search(path.stem)
        return match.group(1).upper() if match else ""

    @staticmethod
    def _category_value_id(qid):
        values = Value.objects.filter(
            concept__legacyoid=f"{PAC_ENTITY_URL}{qid}",
            valuetype_id="prefLabel",
        ).order_by("language_id")
        value = values.filter(language_id="en").first() or values.first()
        if value is None:
            raise CommandError(f"Local PAC pottery category value {qid} was not found.")
        return str(value.valueid)
