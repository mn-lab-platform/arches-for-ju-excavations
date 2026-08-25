"""Import historic pottery workbooks with Project/Trench/Context columns."""

from collections import Counter
from pathlib import Path
import re

from django.core.management.base import CommandError
from django.db import transaction
from openpyxl import load_workbook

from arches.app.models.models import Resource, TileModel, Value

from arches_for_excavation.management.commands.import_pottery_records_excel import Command as RecordsCommand
from arches_for_excavation.utils.pottery.concept_lookup import (
    apply_dictionary_alias,
    get_node_datatype,
    resolve_dictionary_value,
)
from arches_for_excavation.utils.pottery.common import clean_cell, localized_string, normalize_tile_value


LAMP_FILE_RE = re.compile(r"(?:^|[_-])L(?:[_-]|$)", re.IGNORECASE)

# Pottery Collection → Pottery Fragments card. Legacy workbooks are either
# Table Ware or Lamp; their narrative Remarks belong to that category tile.
POTTERY_FRAGMENT_NODEGROUP_ID = "8f7a5ca4-9c49-405d-9a08-a8debb13a9ec"
POTTERY_TYPE_NODE_ID = "3bc235a3-2240-4e94-b8af-f4c70ee13af0"
CATEGORY_REMARKS_NODE_ID = "3c371503-9028-464a-8b85-53a43c853781"
PAC_ENTITY_URL = "https://pac.cenagis.edu.pl/entity/"
RECORD_TYPE_CATEGORY_QIDS = {"table-ware": "Q937", "lamp": "Q924"}
CORE_HEADERS = {
    "Project", "Trench", "Context", "p_no", "SUBCATEGORY", "FORM",
    "VESSEL_PART", "CHRONOLOGY", "PROVENANCE",
}
GENERATED_REPORT_FILENAMES = {
    "missing_pottery_record_collections.xlsx", "pottery_record_context_issues.xlsx",
    "unknown_pottery_record_concepts.xlsx", "unknown_legacy_pottery_record_concepts.xlsx",
}
# Authors declared in TW_Autorstwo_2026_08_10. Keys are case-folded so the
# match remains stable if a spreadsheet is copied with different capitalization.
AUTHORS_BY_FILENAME = {
    "mal_tt_iii_l.xlsx": "Małgorzata Kajzer",
    "mal_tt_iii_tw.xlsx": "Małgorzata Kajzer, Kamila Niziołek",
    "mal_tt_x_r1.xlsx": "Małgorzata Kajzer, Kamila Niziołek",
    "pap_red_slip_jugs.xlsx": "Małgorzata Kajzer",
    "pap_thin-walled_ware.xlsx": "Małgorzata Kajzer",
    "pap_tr_iv_kiln fill.xlsx": "Małgorzata Kajzer",
    "pap_tr_iv_r30.xlsx": "Małgorzata Kajzer, Kamila Niziołek",
    "pap_tr_iv_r31.xlsx": "Małgorzata Kajzer, Kamila Niziołek",
    "pap17_s.117.xlsx": "Małgorzata Kajzer",
}

# Only narrative notes belong in the Pottery Record Comment field.  The
# remaining spreadsheet columns map to their own model fields and must not be
# duplicated into one large audit comment.
EXTRA_COMMENT_HEADERS = ("COMMENTS",)


class Command(RecordsCommand):
    help = "Import historic pottery sheets linked to existing Pottery Collections by full Context."

    def add_arguments(self, parser):
        source = parser.add_mutually_exclusive_group(required=True)
        source.add_argument("--directory", help="Directory containing historic pottery .xlsx files.")
        source.add_argument("--file", action="append", dest="files", help="One historic .xlsx file; may be repeated.")
        parser.add_argument("--apply", action="store_true", help="Create records. Without this flag the command is a dry-run.")
        parser.add_argument(
            "--create-missing-collections",
            action="store_true",
            help="Create one minimal Pottery Collection for each existing Context that has none.",
        )
        parser.add_argument("--dataset", default="legacy-pottery-records", help="Stable prefix used for idempotent source row IDs.")
        parser.add_argument("--missing-collections-report", help="Optional .xlsx path for Contexts without a usable Pottery Collection.")
        parser.add_argument("--audit-unknown-concepts", action="store_true", help="Report unresolved dictionary values without creating records.")
        parser.add_argument("--audit-contexts", action="store_true", help="Report Context and collection-link problems without creating records.")
        parser.add_argument("--context-issues-report", help="Optional .xlsx path for --audit-contexts output.")
        parser.add_argument("--unknown-concepts-report", help="Optional .xlsx path for the unknown-value report.")

    def handle(self, *args, **options):
        if options["apply"] and (options["audit_unknown_concepts"] or options["audit_contexts"]):
            raise CommandError("--apply cannot be used with --audit-unknown-concepts or --audit-contexts.")
        if options["create_missing_collections"] and options["audit_contexts"]:
            raise CommandError("--create-missing-collections cannot be used with --audit-contexts.")
        self.apply = options["apply"]
        self.create_missing_collections = options["create_missing_collections"]
        self.audit_unknown_concepts = options["audit_unknown_concepts"]
        self.audit_contexts = options["audit_contexts"]
        self.unknown_concept_values = {}
        self.dataset = clean_cell(options["dataset"]) or "legacy-pottery-records"
        self._initialize_import_state()
        files = self._source_files(options)
        if not files:
            raise CommandError("No historic pottery .xlsx files found.")
        self.missing_collections_report_path = self._context_issues_report_path(options, files)
        self.unknown_concepts_report_path = self._unknown_concepts_report_path(options, files)
        totals = Counter()
        self.stdout.write(f"Legacy pottery records import [{'APPLY' if self.apply else 'DRY-RUN'}]")
        self.stdout.write(f"Workbooks: {len(files)}")
        for path in files:
            totals.update(self._process_workbook(path, self._record_type_from_path(path)))
        self.stdout.write("Summary:")
        for key in (
            "workbooks_recognized", "workbooks_skipped", "rows_seen", "created",
            "would_create", "unchanged", "skipped_empty", "skipped_red_context",
            "missing_collections", "duplicate_collections", "invalid_contexts",
            "invalid_quantity", "unknown_concepts", "unknown_special_finds", "valid_contexts",
            "collections_would_create", "collections_created",
            "category_remarks_would_update", "category_remarks_updated",
            "category_fragments_would_create", "category_fragments_created", "errors",
        ):
            self.stdout.write(f"  {key}: {totals[key]}")
        if not self.apply:
            if not self.audit_contexts:
                self._write_missing_collections()
            self._write_missing_collection_report()
            if self.audit_contexts:
                self.stdout.write("Context audit only. No records were created.")
            elif self.audit_unknown_concepts:
                self._write_unknown_concepts_report()
                self._write_unknown_concepts()
                self.stdout.write("Unknown-concepts audit only. No records were created.")
            else:
                self.stdout.write("Dry-run only. Add --apply to create records.")

    def _initialize_import_state(self):
        from arches_for_excavation.management.commands.import_collection_excel import Command as CollectionImportCommand
        self.context_importer = CollectionImportCommand()
        self.context_importer.apply = False
        self.context_importer.create_missing_contexts = False
        self.context_importer.context_resource_id = None
        self.context_importer.context_map = {}
        self.context_importer.created_contexts = {}
        self.context_importer.trench_resource_ids = None
        self.collection_ids_by_context = {}
        self.created_collection_ids = {}
        self.missing_collections = {}
        self.category_value_ids = {}

    @staticmethod
    def _unknown_concepts_report_path(options, files):
        explicit_path = clean_cell(options.get("unknown_concepts_report"))
        if explicit_path:
            return Path(explicit_path)
        if options.get("directory"):
            return Path(options["directory"]) / "unknown_legacy_pottery_record_concepts.xlsx"
        return files[0].parent / "unknown_legacy_pottery_record_concepts.xlsx"

    @staticmethod
    def _source_files(options):
        if options.get("directory"):
            directory = Path(options["directory"])
            if not directory.is_dir():
                raise CommandError(f"Directory not found: {directory}")
            return [path for path in sorted(directory.rglob("*.xlsx")) if not path.name.startswith("~$") and path.name not in GENERATED_REPORT_FILENAMES]
        paths = [Path(value) for value in options.get("files", [])]
        missing = [str(path) for path in paths if not path.is_file()]
        if missing:
            raise CommandError(f"Workbook not found: {', '.join(missing)}")
        return paths

    @staticmethod
    def _record_type_from_path(path):
        return "lamp" if LAMP_FILE_RE.search(path.stem) else "table-ware"

    @staticmethod
    def _author_for_path(path):
        return AUTHORS_BY_FILENAME.get(path.name.casefold(), "")

    def _process_workbook(self, path, record_type):
        totals = Counter()
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows = sheet.iter_rows(values_only=False)
                try:
                    header_cells = next(rows)
                except StopIteration:
                    continue
                headers = [clean_cell(cell.value) for cell in header_cells]
                if not CORE_HEADERS.issubset(headers):
                    continue
                totals["workbooks_recognized"] += 1
                context_index = headers.index("Context")
                self.stdout.write(f"Workbook: {path.name}; sheet: {sheet.title} ({record_type})")
                for row_number, cells in enumerate(rows, start=2):
                    if self._is_marked_red(cells[context_index]):
                        source = f"{path.name}:{sheet.title}:{row_number}"
                        context_label = clean_cell(cells[context_index].value) or "<empty>"
                        self._record_missing_collection(context_label, record_type, source, "marked red (excluded)")
                        totals["skipped_red_context"] += 1
                        self.stdout.write(f"  {source}: skipped Context marked red")
                        continue
                    source_row = dict(zip(headers, (cell.value for cell in cells)))
                    # Excel's used range often includes trailing blank rows and
                    # sheet notes. A pottery record needs at least one of these
                    # identifying fields; otherwise it is not a data row.
                    if not any(clean_cell(source_row.get(header)) for header in ("Project", "Trench", "Context", "p_no")):
                        totals["skipped_empty"] += 1
                        continue
                    row, error = self._standard_row(source_row, path, sheet.title, row_number)
                    if error:
                        totals["invalid_contexts"] += 1
                        source = f"{path.name}:{sheet.title}:{row_number}"
                        self._record_missing_collection(clean_cell(source_row.get("Context")) or "<empty>", record_type, source, error)
                        if not self.audit_contexts:
                            self.stderr.write(f"  {source}: {error}")
                        continue
                    totals.update(self._process_row(path, sheet.title, row_number, row, record_type))
        finally:
            workbook.close()
        if not totals["workbooks_recognized"]:
            totals["workbooks_skipped"] += 1
            self.stderr.write(f"Skipped {path.name}: no historic pottery table found")
        return totals

    def _standard_row(self, source, path, sheet_title, row_number):
        context_label, error = self._context_label(source)
        if error:
            return {}, error
        return {
            # Do not fabricate a Form ID. When the historical workbook has no
            # Form ID/form_no value, the Pottery Record must keep it blank.
            "Form ID": clean_cell(
                source.get("Form ID") or source.get("Form_no") or source.get("form_no")
            ),
            "Pottery Collection": context_label,
            "P number": source.get("p_no"),
            "Quantity": source.get("count"),
            "TW MG no": source.get("TW_MG_no"),
            "Author": self._author_for_path(path),
            "Sub-Category": self._non_uncertainty_value(source.get("SUBCATEGORY")),
            "Form": self._non_uncertainty_value(source.get("FORM")),
            "Type": source.get("TYPE") or source.get("TYPE "),
            "Vessel Part": source.get("VESSEL_PART"),
            "Morphology": self._non_uncertainty_value(source.get("Morphology")),
            "Surface Treatment": self._non_uncertainty_value(source.get("Surface treatment")),
            "Period": source.get("CHRONOLOGY"),
            "Provenance": source.get("PROVENANCE"),
            "Drawing": source.get("DRAWN") if "DRAWN" in source else source.get("Drawing"),
            "Photo": source.get("PHOTO") if "PHOTO" in source else source.get("Photo"),
            "Comment": self._comment(source),
            "Category Remarks": clean_cell(source.get("Remarks")),
        }, ""

    def _process_row(self, path, sheet_name, row_number, row, record_type):
        """Create/update the record, then append source Remarks to its collection."""
        totals = super()._process_row(path, sheet_name, row_number, row, record_type)
        if self.audit_contexts or self.audit_unknown_concepts:
            return totals
        if any(totals[key] for key in (
            "errors", "invalid_contexts", "missing_collections", "duplicate_collections", "skipped_empty",
        )):
            return totals
        remarks = clean_cell(row.get("Category Remarks"))
        if not remarks:
            return totals
        context_label = self._value(row, "context")
        collection_ids, context_error = self._collection_ids(context_label)
        if context_error or len(collection_ids) != 1:
            return totals
        totals.update(self._append_category_remarks(
            str(collection_ids[0]), record_type, remarks,
            f"{path.name}:{sheet_name}:{row_number}",
        ))
        return totals

    def _category_value_id(self, record_type):
        if record_type in self.category_value_ids:
            return self.category_value_ids[record_type]
        qid = RECORD_TYPE_CATEGORY_QIDS[record_type]
        values = Value.objects.filter(
            concept__legacyoid=f"{PAC_ENTITY_URL}{qid}",
            valuetype_id="prefLabel",
        ).order_by("language_id")
        value = values.filter(language_id="en").first() or values.first()
        if value is None:
            raise CommandError(f"Local pottery category value {qid} was not found.")
        value_id = str(value.valueid)
        self.category_value_ids[record_type] = value_id
        return value_id

    def _append_category_remarks(self, collection_id, record_type, remarks, source):
        totals = Counter()
        category_value_id = self._category_value_id(record_type)
        fragments = list(TileModel.objects.filter(
            resourceinstance_id=collection_id,
            nodegroup_id=POTTERY_FRAGMENT_NODEGROUP_ID,
            data__contains={POTTERY_TYPE_NODE_ID: category_value_id},
        ))
        if len(fragments) > 1:
            totals["errors"] += 1
            self.stderr.write(
                f"  {source}: Pottery Collection has {len(fragments)} {record_type} fragments; Remarks not changed"
            )
            return totals

        if fragments:
            fragment = fragments[0]
            existing = clean_cell(normalize_tile_value(fragment.data.get(CATEGORY_REMARKS_NODE_ID)))
            existing_lines = {line.strip() for line in existing.splitlines() if line.strip()}
            # Remarks repeated in several P-number rows are written only once.
            if remarks in existing_lines:
                return totals
            new_remarks = f"{existing}\n{remarks}" if existing else remarks
            if self.apply:
                with transaction.atomic():
                    fragment.data = dict(fragment.data or {})
                    fragment.data[CATEGORY_REMARKS_NODE_ID] = localized_string(new_remarks)
                    fragment.save()
                self.stdout.write(f"  {source}: appended category Remarks ({record_type})")
                totals["category_remarks_updated"] += 1
            else:
                self.stdout.write(f"  {source}: would append category Remarks ({record_type})")
                totals["category_remarks_would_update"] += 1
            return totals

        if self.apply:
            with transaction.atomic():
                resource = Resource.objects.get(resourceinstanceid=collection_id)
                self._save_group(resource, {
                    POTTERY_TYPE_NODE_ID: category_value_id,
                    CATEGORY_REMARKS_NODE_ID: localized_string(remarks),
                })
            self.stdout.write(f"  {source}: created {record_type} fragment with category Remarks")
            totals["category_fragments_created"] += 1
        else:
            self.stdout.write(f"  {source}: would create {record_type} fragment with category Remarks")
            totals["category_fragments_would_create"] += 1
        return totals

    def _concept_values(self, key, raw_value, dictionary, node_id, config, source, totals):
        raw_value = clean_cell(raw_value)
        if key == "type" and get_node_datatype(node_id) == "concept-list" and "/" in raw_value:
            # A slash normally separates two independent Types. Some Type
            # names themselves contain a slash (e.g. Hayes P37/X38), so use
            # the split only when every part is a dictionary term.
            parts = [
                self._non_uncertainty_value(part)
                for part in re.split(r"\s*/\s*", raw_value)
                if self._non_uncertainty_value(part)
            ]
            part_ids = [
                resolve_dictionary_value(
                    dictionary,
                    apply_dictionary_alias(key, part, config.get("dictionary_aliases", {})),
                )
                for part in parts
            ]
            if parts and all(part_ids):
                return list(dict.fromkeys(part_ids))

            full_value = self._non_uncertainty_value(raw_value)
            full_id = resolve_dictionary_value(
                dictionary,
                apply_dictionary_alias(key, full_value, config.get("dictionary_aliases", {})),
            )
            if full_id:
                return [full_id]

            totals["unknown_concepts"] += 1
            self._record_unknown(key, full_value, source)
            return []

        if get_node_datatype(node_id) == "concept-list" and key != "type":
            separator = r"\s*(?:\+|/|\?\s+)\s*" if key == "vesselPart" else r"\s*(?:\+|/)\s*"
            values = [value for value in re.split(separator, raw_value) if value]
            resolved = []
            for value in values:
                resolved.extend(super()._concept_values(
                    key, self._non_uncertainty_value(value), dictionary, node_id, config, source, totals,
                ))
            return list(dict.fromkeys(resolved))
        return super()._concept_values(key, raw_value, dictionary, node_id, config, source, totals)

    @staticmethod
    def _non_uncertainty_value(value):
        """Strip question marks from legacy non-uncertainty dictionary fields."""
        return clean_cell(value).replace("?", "").strip()

    @staticmethod
    def _context_label(source):
        project = clean_cell(source.get("Project")).upper()
        trench = clean_cell(source.get("Trench")).upper()
        raw_context = clean_cell(source.get("Context"))
        if not project or not trench or not raw_context:
            return "", "missing Project, Trench or Context"
        try:
            context_number = str(int(float(raw_context)))
        except ValueError:
            return "", f"invalid Context {raw_context!r}"
        trench = re.sub(r"[.\s_-]+", "_", trench).strip("_")
        trench_name = f"{project}_{trench}" if trench.startswith(("TT_", "T_")) else f"{project}_T_{trench}"
        return f"{trench_name}-{context_number}", ""

    @staticmethod
    def _comment(source):
        return "\n".join(
            f"{header}: {value}" for header in EXTRA_COMMENT_HEADERS
            if (value := clean_cell(source.get(header)))
        )
